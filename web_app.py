#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
小型氧化加工厂财务管理系统 - Web版
==================================
基于 Flask 的现代化 Web 界面

依赖安装：
    pip install flask openpyxl

启动：
    python web_app.py
    或双击 启动WEB系统.bat
"""
import os, sys, shutil, glob, csv, io
from datetime import datetime, date
from flask import Flask, render_template, request, redirect, url_for, flash, jsonify

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from 氧化加工厂财务系统 import (
    DataManager, EXPENSE_CATEGORIES, MATERIALS, SAMPLE_CUSTOMERS,
    VERSION, fmt, calc_income_tax, EXCEL_FILE, DATA_DIR
)

BACKUP_DIR = os.path.join(DATA_DIR, 'backups')
os.makedirs(BACKUP_DIR, exist_ok=True)

app = Flask(__name__)
app.secret_key = 'caiwu_system_2026'
app.jinja_env.globals.update(today=lambda: date.today())

def get_dm():
    dm = DataManager()
    return dm

def get_income_data(dm):
    ws = dm.sheet("收入记录")
    rows = dm.read_rows(ws, start_row=4, end_col=10)
    data = []
    for r in rows:
        if r[0] and r[2]:
            data.append({
                "date": r[0], "customer": str(r[2] or ""),
                "amount": to_num(r[3]),
                "method": str(r[4] or ""), "tax_type": str(r[5] or ""),
                "invoice": str(r[9] or ""), "note": ""
            })
    return data

def get_expense_data(dm):
    ws = dm.sheet("支出记录")
    rows = dm.read_rows(ws, start_row=4, end_col=7)
    data = []
    for r in rows:
        if r[0] and r[3]:
            data.append({
                "date": r[0], "category": str(r[2] or ""),
                "amount": to_num(r[3]),
                "supplier": str(r[4] or ""), "method": str(r[5] or ""),
                "note": str(r[6] or "")
            })
    return data

def get_arap_data(dm):
    ws = dm.sheet("应收应付")
    rows = dm.read_rows(ws, start_row=4, end_col=7)
    data = []
    for r in rows:
        if r[0]:
            opening = to_num(r[1])
            increase = to_num(r[2])
            receipt = to_num(r[3])
            offset = to_num(r[4])
            ending = to_num(r[5])
            data.append({
                "customer": str(r[0] or ""), "type": "应收",
                "opening": opening, "increase": increase,
                "receipt": receipt, "offset": offset,
                "ending": ending, "note": str(r[6] or ""),
                "balance": ending
            })
    return data

def get_salary_data(dm):
    ws = dm.sheet("工资表")
    rows = dm.read_rows(ws, start_row=5, end_col=10)
    data = []
    for r in rows:
        if r[1]:
            data.append({
                "name": str(r[1]), "base": to_num(r[2]),
                "overtime": to_num(r[3]), "bonus": to_num(r[4]),
                "social": to_num(r[6]), "tax": to_num(r[8]),
                "net": to_num(r[9])
            })
    return data

def get_material_data(dm):
    ws = dm.sheet("材料进销存台账")
    rows = dm.read_rows(ws, start_row=4, end_col=10)
    data = []
    for r in rows:
        if r[1]:
            data.append({
                "date": r[0], "name": str(r[1] or ""), "spec": str(r[2] or ""),
                "qty": to_num(r[3]), "price": to_num(r[4]),
                "amount": to_num(r[5])
            })
    return data

def to_num(v):
    if v is None: return 0
    if isinstance(v, (int, float)): return v
    try: return float(str(v).replace(',', '').replace('¥', '').replace(' ', ''))
    except: return 0

def calc_summary(dm):
    incomes = get_income_data(dm)
    expenses = get_expense_data(dm)
    arap = get_arap_data(dm)
    salaries = get_salary_data(dm)
    materials = get_material_data(dm)

    total_inc = sum(to_num(i["amount"]) for i in incomes)
    total_exp = sum(to_num(e["amount"]) for e in expenses)
    profit = total_inc - total_exp
    profit_rate = (profit / total_inc * 100) if total_inc else 0
    total_ar = sum(to_num(a["balance"]) for a in arap if a["type"] == "应收")
    total_ap = sum(to_num(a["balance"]) for a in arap if a["type"] == "应付")

    return {
        "inc_count": len(incomes), "exp_count": len(expenses),
        "ar_count": len(arap), "sal_count": len(salaries),
        "mat_count": len(materials),
        "total_inc": total_inc, "total_exp": total_exp,
        "profit": profit, "profit_rate": profit_rate,
        "total_ar": total_ar, "total_ap": total_ap
    }

@app.route('/')
def dashboard():
    dm = get_dm()
    s = calc_summary(dm)
    return render_template('dashboard.html', summary=s, version=VERSION)

@app.route('/income')
def income():
    dm = get_dm()
    data = get_income_data(dm)
    return render_template('income.html', data=data, version=VERSION)

@app.route('/income/add', methods=['POST'])
def income_add():
    dm = get_dm()
    ws = dm.sheet("收入记录")
    row = dm.next_row(ws)
    dm.write_row(ws, row, [
        request.form.get('date', str(date.today())),
        request.form.get('customer', ''),
        float(request.form.get('amount', 0)),
        request.form.get('method', '银行转账'),
        request.form.get('tax_type', '含税'),
        request.form.get('invoice', '否'),
        request.form.get('note', '')
    ])
    dm.save()
    flash('收入已添加', 'success')
    return redirect(url_for('income'))

@app.route('/api/income/batch_add', methods=['POST'])
def api_income_batch_add():
    dm = get_dm()
    ws = dm.sheet("收入记录")
    items = request.json
    count = 0
    for item in items:
        row = dm.next_row(ws)
        dm.write_row(ws, row, [
            item.get('date', str(date.today())),
            '',
            item.get('customer', ''),
            float(item.get('amount', 0)),
            item.get('method', '银行转账'),
            item.get('tax_type', '含税'),
            '', '',
            '',
            item.get('invoice', '否')
        ])
        count += 1
    dm.save()
    return jsonify({"ok": True, "count": count})

@app.route('/api/income/import_csv', methods=['POST'])
def api_income_import_csv():
    dm = get_dm()
    if 'file' not in request.files:
        return jsonify({"ok": False, "error": "未上传文件"}), 400
    f = request.files['file']
    if not f.filename.endswith('.csv'):
        return jsonify({"ok": False, "error": "请上传CSV文件"}), 400

    content = f.read().decode('utf-8-sig')
    reader = csv.reader(io.StringIO(content))
    ws = dm.sheet("收入记录")
    count = 0
    for row_data in reader:
        if not row_data or not row_data[0].strip():
            continue
        if row_data[0].strip() in ('日期', 'date'):
            continue
        date_val = row_data[0].strip()
        customer = row_data[2].strip() if len(row_data) > 2 else ''
        amount = float(row_data[3]) if len(row_data) > 3 and row_data[3].strip() else 0
        method = row_data[4].strip() if len(row_data) > 4 else '银行转账'
        tax_type = row_data[5].strip() if len(row_data) > 5 else '含税'
        invoice = row_data[9].strip() if len(row_data) > 9 else '否'
        r = dm.next_row(ws)
        dm.write_row(ws, r, [date_val, '', customer, amount, method, tax_type, '', '', '', invoice])
        count += 1
    dm.save()
    return jsonify({"ok": True, "count": count})

@app.route('/export/income_template')
def export_income_template():
    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow(['日期', '凭证号', '客户名称', '收入金额', '收款方式', '含税/不含税', '税率', '税额', '不含税金额', '开票状态'])
    w.writerow(['2026-06-01', '', '华鑫铝业', '15000', '银行转账', '含税', '', '', '', '未开票'])
    w.writerow(['2026-06-02', '', '永达五金', '8000', '微信', '含税', '', '', '', '未开票'])
    out = io.BytesIO(buf.getvalue().encode('utf-8-sig'))
    return out.read(), 200, {
        'Content-Type': 'text/csv; charset=utf-8-sig',
        'Content-Disposition': 'attachment; filename="收入导入模板.csv"'
    }

@app.route('/expense')
def expense():
    dm = get_dm()
    data = get_expense_data(dm)
    return render_template('expense.html', data=data, categories=EXPENSE_CATEGORIES, version=VERSION)

@app.route('/expense/add', methods=['POST'])
def expense_add():
    dm = get_dm()
    ws = dm.sheet("支出记录")
    row = dm.next_row(ws)
    dm.write_row(ws, row, [
        request.form.get('date', str(date.today())),
        request.form.get('category', ''),
        float(request.form.get('amount', 0)),
        request.form.get('supplier', ''),
        request.form.get('method', '银行转账'),
        request.form.get('note', '')
    ])
    dm.save()
    flash('支出已添加', 'success')
    return redirect(url_for('expense'))

@app.route('/api/expense/batch_add', methods=['POST'])
def api_expense_batch_add():
    dm = get_dm()
    ws = dm.sheet("支出记录")
    items = request.json
    count = 0
    for item in items:
        row = dm.next_row(ws)
        dm.write_row(ws, row, [
            item.get('date', str(date.today())),
            '',
            item.get('category', ''),
            float(item.get('amount', 0)),
            item.get('supplier', ''),
            item.get('method', '银行转账'),
            item.get('note', '')
        ])
        count += 1
    dm.save()
    return jsonify({"ok": True, "count": count})

@app.route('/api/expense/import_csv', methods=['POST'])
def api_expense_import_csv():
    dm = get_dm()
    if 'file' not in request.files:
        return jsonify({"ok": False, "error": "未上传文件"}), 400
    f = request.files['file']
    if not f.filename.endswith('.csv'):
        return jsonify({"ok": False, "error": "请上传CSV文件"}), 400

    content = f.read().decode('utf-8-sig')
    reader = csv.reader(io.StringIO(content))
    ws = dm.sheet("支出记录")
    count = 0
    for row_data in reader:
        if not row_data or not row_data[0].strip():
            continue
        if row_data[0].strip() in ('日期', 'date'):
            continue
        date_val = row_data[0].strip()
        category = row_data[2].strip() if len(row_data) > 2 else ''
        amount = float(row_data[3]) if len(row_data) > 3 and row_data[3].strip() else 0
        supplier = row_data[4].strip() if len(row_data) > 4 else ''
        method = row_data[5].strip() if len(row_data) > 5 else '银行转账'
        note = row_data[6].strip() if len(row_data) > 6 else ''
        r = dm.next_row(ws)
        dm.write_row(ws, r, [date_val, '', category, amount, supplier, method, note])
        count += 1
    dm.save()
    return jsonify({"ok": True, "count": count})

@app.route('/export/expense_template')
def export_expense_template():
    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow(['日期', '凭证号', '类别', '金额', '供应商', '付款方式', '备注'])
    w.writerow(['2026-06-01', '', '厂租', '8000', '房东', '银行转账', '6月厂租'])
    w.writerow(['2026-06-02', '', '原材料', '5000', '化工公司', '银行转账', ''])
    out = io.BytesIO(buf.getvalue().encode('utf-8-sig'))
    return out.read(), 200, {
        'Content-Type': 'text/csv; charset=utf-8-sig',
        'Content-Disposition': 'attachment; filename="支出导入模板.csv"'
    }

@app.route('/arap')
def arap():
    dm = get_dm()
    data = get_arap_data(dm)
    return render_template('arap.html', data=data, version=VERSION)

@app.route('/report/profit')
def profit_report():
    dm = get_dm()
    s = calc_summary(dm)
    expenses = get_expense_data(dm)
    by_category = {}
    for e in expenses:
        cat = e["category"] or "其他"
        by_category[cat] = by_category.get(cat, 0) + (e["amount"] or 0)
    cat_list = sorted(by_category.items(), key=lambda x: -x[1])
    return render_template('profit.html', summary=s, categories=cat_list, version=VERSION)

@app.route('/report/balance')
def balance_report():
    dm = get_dm()
    s = calc_summary(dm)
    return render_template('balance.html', summary=s, version=VERSION)

@app.route('/report/monthly')
def monthly_report():
    dm = get_dm()
    incomes = get_income_data(dm)
    expenses = get_expense_data(dm)
    monthly = {}
    for i in incomes:
        d = i["date"]
        if d and hasattr(d, 'month'):
            key = f"{d.year}-{d.month:02d}"
            monthly.setdefault(key, {"收入": 0, "支出": 0})
            monthly[key]["收入"] += i["amount"] or 0
    for e in expenses:
        d = e["date"]
        if d and hasattr(d, 'month'):
            key = f"{d.year}-{d.month:02d}"
            monthly.setdefault(key, {"收入": 0, "支出": 0})
            monthly[key]["支出"] += e["amount"] or 0

    months = sorted(monthly.keys())
    chart_labels = months
    chart_income = [monthly[m]["收入"] for m in months]
    chart_expense = [monthly[m]["支出"] for m in months]
    chart_profit = [monthly[m]["收入"] - monthly[m]["支出"] for m in months]

    return render_template('monthly.html',
        months=months, monthly=monthly,
        labels=chart_labels, income_data=chart_income,
        expense_data=chart_expense, profit_data=chart_profit,
        version=VERSION
    )

@app.route('/salary')
def salary():
    dm = get_dm()
    data = get_salary_data(dm)
    return render_template('salary.html', data=data, version=VERSION)

@app.route('/api/salary/batch_add', methods=['POST'])
def api_salary_batch_add():
    dm = get_dm()
    ws = dm.sheet("工资表")
    items = request.json
    count = 0
    for idx, item in enumerate(items):
        row = dm.next_row(ws)
        base = float(item.get('base', 0))
        overtime = float(item.get('overtime', 0))
        bonus = float(item.get('bonus', 0))
        gross = base + overtime + bonus
        social = float(item.get('social', gross * 0.1))
        accfund = float(item.get('accfund', 0))
        tax = float(item.get('tax', 0))
        net = gross - social - accfund - tax
        dm.write_row(ws, row, [
            idx + 1, item.get('name', ''),
            base, overtime, bonus, gross,
            social, accfund, tax, net
        ])
        count += 1
    dm.save()
    return jsonify({"ok": True, "count": count})

@app.route('/api/salary/delete', methods=['POST'])
def api_salary_delete():
    dm = get_dm()
    idx = request.json.get('index', -1)
    ws = dm.sheet("工资表")
    row = idx + 5
    if 5 <= row <= dm.last_row(ws):
        ws.delete_rows(row)
        dm.save()
        return jsonify({"ok": True})
    return jsonify({"ok": False, "error": "无效行号"}), 400

@app.route('/api/salary/import_csv', methods=['POST'])
def api_salary_import_csv():
    dm = get_dm()
    if 'file' not in request.files:
        return jsonify({"ok": False, "error": "未上传文件"}), 400
    f = request.files['file']
    if not f.filename.endswith('.csv'):
        return jsonify({"ok": False, "error": "请上传CSV文件"}), 400
    content = f.read().decode('utf-8-sig')
    reader = csv.reader(io.StringIO(content))
    ws = dm.sheet("工资表")
    count = 0
    for row_data in reader:
        if not row_data or not row_data[1].strip():
            continue
        if row_data[1].strip() in ('姓名', 'name'):
            continue
        name = row_data[1].strip()
        base = float(row_data[2]) if len(row_data) > 2 and row_data[2].strip() else 0
        overtime = float(row_data[3]) if len(row_data) > 3 and row_data[3].strip() else 0
        bonus = float(row_data[4]) if len(row_data) > 4 and row_data[4].strip() else 0
        gross = base + overtime + bonus
        social = float(row_data[6]) if len(row_data) > 6 and row_data[6].strip() else round(gross * 0.1, 2)
        accfund = float(row_data[7]) if len(row_data) > 7 and row_data[7].strip() else 0
        tax = float(row_data[8]) if len(row_data) > 8 and row_data[8].strip() else 0
        net = gross - social - accfund - tax
        r = dm.next_row(ws)
        dm.write_row(ws, r, [count + 1, name, base, overtime, bonus, gross, social, accfund, tax, net])
        count += 1
    dm.save()
    return jsonify({"ok": True, "count": count})

@app.route('/export/salary_template')
def export_salary_template():
    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow(['序号', '姓名', '基本工资', '加班费', '奖金', '应发合计', '社保个人', '公积金个人', '个税', '实发合计'])
    w.writerow([1, '张三', 3500, 560, 238, 4298, 430, 0, 0, 3868])
    w.writerow([2, '李四', 4200, 300, 500, 5000, 500, 0, 0, 4500])
    out = io.BytesIO(buf.getvalue().encode('utf-8-sig'))
    return out.read(), 200, {
        'Content-Type': 'text/csv; charset=utf-8-sig',
        'Content-Disposition': 'attachment; filename="工资导入模板.csv"'
    }

@app.route('/material')
def material():
    dm = get_dm()
    data = get_material_data(dm)
    return render_template('material.html', data=data, materials=MATERIALS, version=VERSION)

@app.route('/api/material/batch_add', methods=['POST'])
def api_material_batch_add():
    dm = get_dm()
    ws = dm.sheet("材料进销存台账")
    items = request.json
    count = 0
    for item in items:
        row = dm.next_row(ws)
        in_qty = float(item.get('in_qty', 0))
        in_price = float(item.get('in_price', 0))
        out_qty = float(item.get('out_qty', 0))
        out_price = float(item.get('out_price', 0))
        in_amount = in_qty * in_price
        out_amount = out_qty * out_price
        dm.write_row(ws, row, [
            item.get('date', str(date.today())),
            item.get('name', ''),
            item.get('spec', ''),
            in_qty if in_qty else None,
            in_price if in_price else None,
            in_amount if in_amount else None,
            out_qty if out_qty else None,
            out_price if out_price else None,
            out_amount if out_amount else None,
            None
        ])
        count += 1
    dm.save()
    return jsonify({"ok": True, "count": count})

@app.route('/api/material/delete', methods=['POST'])
def api_material_delete():
    dm = get_dm()
    idx = request.json.get('index', -1)
    ws = dm.sheet("材料进销存台账")
    row = idx + 4
    if 4 <= row <= dm.last_row(ws):
        ws.delete_rows(row)
        dm.save()
        return jsonify({"ok": True})
    return jsonify({"ok": False, "error": "无效行号"}), 400

@app.route('/api/material/import_csv', methods=['POST'])
def api_material_import_csv():
    dm = get_dm()
    if 'file' not in request.files:
        return jsonify({"ok": False, "error": "未上传文件"}), 400
    f = request.files['file']
    if not f.filename.endswith('.csv'):
        return jsonify({"ok": False, "error": "请上传CSV文件"}), 400
    content = f.read().decode('utf-8-sig')
    reader = csv.reader(io.StringIO(content))
    ws = dm.sheet("材料进销存台账")
    count = 0
    for row_data in reader:
        if not row_data or not row_data[1].strip():
            continue
        if row_data[1].strip() in ('材料名称', 'name'):
            continue
        date_val = row_data[0].strip()
        name = row_data[1].strip()
        spec = row_data[2].strip() if len(row_data) > 2 else ''
        in_qty = float(row_data[3]) if len(row_data) > 3 and row_data[3].strip() else 0
        in_price = float(row_data[4]) if len(row_data) > 4 and row_data[4].strip() else 0
        out_qty = float(row_data[6]) if len(row_data) > 6 and row_data[6].strip() else 0
        out_price = float(row_data[7]) if len(row_data) > 7 and row_data[7].strip() else 0
        in_amount = in_qty * in_price
        out_amount = out_qty * out_price
        r = dm.next_row(ws)
        dm.write_row(ws, r, [
            date_val, name, spec,
            in_qty if in_qty else None, in_price if in_price else None, in_amount if in_amount else None,
            out_qty if out_qty else None, out_price if out_price else None, out_amount if out_amount else None,
            None
        ])
        count += 1
    dm.save()
    return jsonify({"ok": True, "count": count})

@app.route('/export/material_template')
def export_material_template():
    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow(['日期', '材料名称', '规格', '入库数量', '入库单价', '入库金额', '出库数量', '出库单价', '出库金额', '结存数量'])
    w.writerow(['2026-06-01', '亚钠', '25kg/袋', 20, 180, 3600, '', '', '', ''])
    w.writerow(['2026-06-02', '片碱', '25kg/袋', '', '', '', 10, 200, 2000, ''])
    out = io.BytesIO(buf.getvalue().encode('utf-8-sig'))
    return out.read(), 200, {
        'Content-Type': 'text/csv; charset=utf-8-sig',
        'Content-Disposition': 'attachment; filename="材料导入模板.csv"'
    }

@app.route('/api/summary')
def api_summary():
    dm = get_dm()
    s = calc_summary(dm)
    return jsonify(s)

@app.route('/api/income')
def api_income():
    dm = get_dm()
    return jsonify(get_income_data(dm))

@app.route('/api/income/search')
def api_income_search():
    dm = get_dm()
    q = request.args.get('q', '').lower()
    data = get_income_data(dm)
    if q:
        data = [d for d in data if q in str(d.get('customer','')).lower() or q in str(d.get('date',''))]
    return jsonify(data)

@app.route('/api/expense/search')
def api_expense_search():
    dm = get_dm()
    q = request.args.get('q', '').lower()
    cat = request.args.get('cat', '')
    data = get_expense_data(dm)
    if q:
        data = [d for d in data if q in str(d.get('category','')).lower() or q in str(d.get('supplier','')).lower()]
    if cat:
        data = [d for d in data if d.get('category') == cat]
    return jsonify(data)

@app.route('/api/income/delete', methods=['POST'])
def api_income_delete():
    dm = get_dm()
    idx = request.json.get('index', -1)
    ws = dm.sheet("收入记录")
    row = idx + 4
    if 4 <= row <= dm.last_row(ws):
        ws.delete_rows(row)
        dm.save()
        return jsonify({"ok": True})
    return jsonify({"ok": False, "error": "无效行号"}), 400

@app.route('/api/expense/delete', methods=['POST'])
def api_expense_delete():
    dm = get_dm()
    idx = request.json.get('index', -1)
    ws = dm.sheet("支出记录")
    row = idx + 4
    if 4 <= row <= dm.last_row(ws):
        ws.delete_rows(row)
        dm.save()
        return jsonify({"ok": True})
    return jsonify({"ok": False, "error": "无效行号"}), 400

@app.route('/api/customers')
def api_customers():
    dm = get_dm()
    ws = dm.sheet("收入记录")
    seen = set()
    customers = []
    for r in dm.read_rows(ws, start_row=4, end_col=3):
        if r[2] and str(r[2]) not in seen:
            seen.add(str(r[2]))
            customers.append(str(r[2]))
    return jsonify(sorted(customers))

@app.route('/api/alerts')
def api_alerts():
    dm = get_dm()
    alerts = []
    arap = get_arap_data(dm)
    for a in arap:
        if a["type"] == "应收" and a["balance"] > 0:
            if a["balance"] > 20000:
                alerts.append({"type": "danger", "icon": "bi-exclamation-triangle", "msg": f"应收 {a['customer']}：¥{a['balance']:,.2f}（大额应收）"})
            elif a["balance"] > 10000:
                alerts.append({"type": "warning", "icon": "bi-exclamation-circle", "msg": f"应收 {a['customer']}：¥{a['balance']:,.2f}"})

    ws = dm.sheet("材料进销存台账")
    stock = {}
    for r in dm.read_rows(ws, start_row=4, end_col=10):
        name = str(r[1] or "")
        in_qty = to_num(r[3]); out_qty = to_num(r[6])
        bal = stock.get(name, 0) + in_qty - out_qty
        stock[name] = bal
    for name, bal in sorted(stock.items()):
        if name and bal < 10:
            alerts.append({"type": "warning" if bal > 0 else "danger", "icon": "bi-box", "msg": f"材料 {name}：库存 {bal}（{'缺货' if bal <= 0 else '低于安全线'}）"})

    return jsonify(alerts[:8])

@app.route('/export/profit')
def export_profit():
    dm = get_dm()
    s = calc_summary(dm)
    expenses = get_expense_data(dm)
    by_cat = {}
    for e in expenses:
        c = e["category"] or "其他"
        by_cat[c] = by_cat.get(c, 0) + (e["amount"] or 0)

    from openpyxl import Workbook
    from io import BytesIO
    wb = Workbook()
    ws = wb.active
    ws.title = "利润报表"
    ws['A1'] = '利润报表'; ws.merge_cells('A1:C1')
    ws['A3'] = '项目'; ws['B3'] = '金额'
    ws['A4'] = '总收入'; ws['B4'] = s['total_inc']
    ws['A5'] = '总支出'; ws['B5'] = s['total_exp']
    ws['A6'] = '净利润'; ws['B6'] = s['profit']
    ws['A8'] = '支出类别'; ws['B8'] = '金额'
    row = 9
    for cat, amt in sorted(by_cat.items(), key=lambda x: -x[1]):
        ws.cell(row=row, column=1, value=cat)
        ws.cell(row=row, column=2, value=amt)
        row += 1

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read(), 200, {
        'Content-Type': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        'Content-Disposition': 'attachment; filename="利润报表.xlsx"'
    }

@app.route('/export/monthly')
def export_monthly():
    dm = get_dm()
    incomes = get_income_data(dm)
    expenses = get_expense_data(dm)
    monthly = {}
    for i in incomes:
        d = i["date"]
        if d:
            ds = str(d) if not hasattr(d, 'strftime') else d.strftime('%Y-%m')
            key = ds[:7]
            monthly.setdefault(key, {"收入": 0, "支出": 0})
            monthly[key]["收入"] += i["amount"] or 0
    for e in expenses:
        d = e["date"]
        if d:
            ds = str(d) if not hasattr(d, 'strftime') else d.strftime('%Y-%m')
            key = ds[:7]
            monthly.setdefault(key, {"收入": 0, "支出": 0})
            monthly[key]["支出"] += e["amount"] or 0

    from openpyxl import Workbook
    from io import BytesIO
    wb = Workbook()
    ws = wb.active
    ws.title = "月度报表"
    ws['A1'] = '月度报表'; ws.merge_cells('A1:D1')
    ws['A3'] = '月份'; ws['B3'] = '收入'; ws['C3'] = '支出'; ws['D3'] = '利润'
    row = 4
    for m in sorted(monthly.keys()):
        r = monthly[m]
        ws.cell(row=row, column=1, value=m)
        ws.cell(row=row, column=2, value=r["收入"])
        ws.cell(row=row, column=3, value=r["支出"])
        ws.cell(row=row, column=4, value=r["收入"] - r["支出"])
        row += 1

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read(), 200, {
        'Content-Type': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        'Content-Disposition': 'attachment; filename="月度报表.xlsx"'
    }

@app.route('/api/arap')
def api_arap():
    dm = get_dm()
    return jsonify(get_arap_data(dm))

@app.route('/api/arap/search')
def api_arap_search():
    dm = get_dm()
    q = request.args.get('q', '').lower()
    data = get_arap_data(dm)
    if q:
        data = [d for d in data if q in d.get('customer', '').lower()]
    return jsonify(data)

@app.route('/api/arap/add', methods=['POST'])
def api_arap_add():
    dm = get_dm()
    ws = dm.sheet("应收应付")
    j = request.json
    if not j or not j.get('customer'):
        return jsonify({"ok": False, "error": "客户名称为空"}), 400

    opening = to_num(j.get('opening'))
    increase = to_num(j.get('increase'))
    receipt = to_num(j.get('receipt'))
    offset = to_num(j.get('offset'))
    ending = opening + increase - receipt - offset

    row = dm.next_row(ws)
    dm.write_row(ws, row, [
        j['customer'], opening, increase, receipt, offset, ending,
        j.get('note', '')
    ])
    dm.save()
    return jsonify({"ok": True, "ending": ending})

@app.route('/api/arap/delete', methods=['POST'])
def api_arap_delete():
    dm = get_dm()
    idx = request.json.get('index', -1)
    ws = dm.sheet("应收应付")
    row = idx + 4
    if 4 <= row <= dm.last_row(ws):
        ws.delete_rows(row)
        dm.save()
        return jsonify({"ok": True})
    return jsonify({"ok": False, "error": "无效行号"}), 400

@app.route('/api/arap/import_csv', methods=['POST'])
def api_arap_import_csv():
    dm = get_dm()
    if 'file' not in request.files:
        return jsonify({"ok": False, "error": "未上传文件"}), 400
    f = request.files['file']
    if not f.filename.endswith('.csv'):
        return jsonify({"ok": False, "error": "请上传CSV文件"}), 400

    import csv, io
    content = f.read().decode('utf-8-sig')
    reader = csv.reader(io.StringIO(content))
    ws = dm.sheet("应收应付")
    count = 0
    for row_data in reader:
        if not row_data or not row_data[0].strip():
            continue
        customer = row_data[0].strip()
        if customer in ('客户名称', '客户名', 'customer'):
            continue
        opening = to_num(row_data[1]) if len(row_data) > 1 else 0
        increase = to_num(row_data[2]) if len(row_data) > 2 else 0
        receipt = to_num(row_data[3]) if len(row_data) > 3 else 0
        offset = to_num(row_data[4]) if len(row_data) > 4 else 0
        ending = opening + increase - receipt - offset
        note = row_data[6].strip() if len(row_data) > 6 else ''
        r = dm.next_row(ws)
        dm.write_row(ws, r, [customer, opening, increase, receipt, offset, ending, note])
        count += 1

    dm.save()
    return jsonify({"ok": True, "count": count})

@app.route('/export/arap_template')
def export_arap_template():
    import csv, io
    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow(['客户名称', '期初应收', '本期增加', '本期收款', '对冲减少', '备注'])
    w.writerow(['示例-华鑫铝业', 5000, 30000, 25000, 0, '期初余额5000，本期增加30000'])
    w.writerow(['示例-永达五金', 3000, 20000, 18000, 0, ''])
    out = io.BytesIO(buf.getvalue().encode('utf-8-sig'))
    return out.read(), 200, {
        'Content-Type': 'text/csv; charset=utf-8-sig',
        'Content-Disposition': 'attachment; filename="应收应付导入模板.csv"'
    }

@app.route('/backup')
def backup():
    return render_template('backup.html', version=VERSION)

@app.route('/api/backup/list')
def api_backup_list():
    files = []
    for f in sorted(glob.glob(os.path.join(BACKUP_DIR, '*.xlsx')), key=os.path.getmtime, reverse=True):
        size = os.path.getsize(f)
        mtime = datetime.fromtimestamp(os.path.getmtime(f))
        files.append({
            "name": os.path.basename(f),
            "size": size,
            "size_str": f"{size/1024:.1f} KB" if size < 1024*1024 else f"{size/1024/1024:.2f} MB",
            "mtime": mtime.strftime('%Y-%m-%d %H:%M:%S')
        })
    return jsonify(files)

@app.route('/api/backup/create', methods=['POST'])
def api_backup_create():
    ts = datetime.now().strftime('%Y%m%d_%H%M%S')
    dst = os.path.join(BACKUP_DIR, f'财务数据_{ts}.xlsx')
    try:
        # 先保存当前数据
        dm = get_dm()
        dm.save()
        shutil.copy2(EXCEL_FILE, dst)
        return jsonify({"ok": True, "file": os.path.basename(dst)})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500

@app.route('/api/backup/restore', methods=['POST'])
def api_backup_restore():
    name = request.json.get('name', '')
    if not name:
        return jsonify({"ok": False, "error": "未指定备份文件"}), 400
    src = os.path.join(BACKUP_DIR, name)
    if not os.path.isfile(src):
        return jsonify({"ok": False, "error": "备份文件不存在"}), 404
    try:
        shutil.copy2(src, EXCEL_FILE)
        return jsonify({"ok": True})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500

@app.route('/api/backup/delete', methods=['POST'])
def api_backup_delete():
    name = request.json.get('name', '')
    if not name:
        return jsonify({"ok": False, "error": "未指定备份文件"}), 400
    fp = os.path.join(BACKUP_DIR, name)
    if not os.path.isfile(fp):
        return jsonify({"ok": False, "error": "备份文件不存在"}), 404
    try:
        os.remove(fp)
        return jsonify({"ok": True})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500

@app.route('/api/system/stats')
def api_system_stats():
    dm = get_dm()
    incomes = get_income_data(dm)
    expenses = get_expense_data(dm)
    arap = get_arap_data(dm)
    salaries = get_salary_data(dm)
    materials = get_material_data(dm)
    return jsonify({
        "income": len(incomes), "expense": len(expenses),
        "arap": len(arap), "salary": len(salaries), "material": len(materials)
    })

@app.route('/api/system/init', methods=['POST'])
def api_system_init():
    try:
        dm = get_dm()
        # 自动备份
        ts = datetime.now().strftime('%Y%m%d_%H%M%S')
        backup_path = os.path.join(BACKUP_DIR, f'初始化前自动备份_{ts}.xlsx')
        dm.save()
        shutil.copy2(EXCEL_FILE, backup_path)

        # 清空各表数据行（保留表头）
        sheets = [
            ("收入记录", 4), ("支出记录", 4),
            ("应收应付", 4), ("材料进销存台账", 4),
            ("工资表", 5)
        ]
        for name, header_row in sheets:
            ws = dm.sheet(name)
            last = dm.last_row(ws)
            if last >= header_row:
                ws.delete_rows(header_row, last - header_row + 1)

        dm.save()
        return jsonify({"ok": True, "backup": os.path.basename(backup_path)})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500

@app.route('/_ah/health')
def health():
    return 'ok'

@app.route('/data-import')
def data_import():
    return render_template('import.html', version=VERSION)

if __name__ == '__main__':
    print(f"  Web版 {VERSION}")
    print(f"  数据文件: {EXCEL_FILE}")
    print(f"  访问地址: http://127.0.0.1:5000")
    print(f"  按 Ctrl+C 停止服务器")
    app.run(debug=True, host='0.0.0.0', port=5000)
