#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
小型氧化加工厂财务管理系统 V18.0 界面优化版
============================================
完整45个功能 + 5个常用模板 + 智能提醒系统
无需WPS/Excel VBA环境，纯Python独立运行

使用方法：
    python 氧化加工厂财务系统.py
    或双击 启动财务系统.bat（Windows）

依赖安装：
    pip install openpyxl

版本：V18.0
更新：2026年5月
"""

import os
import sys
import csv
import shutil
import random
from datetime import datetime, date, timedelta
from typing import Dict, List, Optional, Tuple, Any

# Windows控制台UTF-8支持
if sys.platform == 'win32':
    import io
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8', errors='replace')

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.formatting.rule import FormulaRule, CellIsRule
except ImportError:
    print("=" * 50)
    print("  缺少依赖库！请运行：")
    print("  pip install openpyxl")
    print("=" * 50)
    sys.exit(1)

# ============================================================================
# 全局配置
# ============================================================================

VERSION = "V18.0 界面优化版"
APP_DIR = os.path.dirname(os.path.abspath(__file__))
DATA_DIR = os.path.join(APP_DIR, "财务数据")
BACKUP_DIR = os.path.join(APP_DIR, "备份")
EXCEL_FILE = os.path.join(DATA_DIR, "财务数据.xlsx")

EXPENSE_CATEGORIES = [
    "厂租", "电费", "水费", "化工", "染料封闭剂",
    "工资", "包装材料", "货车运输费", "外发加工费",
    "税金", "社保费", "业务费", "办公用品费", "维修费", "挂具费"
]

MATERIALS = ["亚钠", "片碱", "硝酸", "磷酸", "硫酸", "封闭剂", "染料", "除油剂"]
SAMPLE_CUSTOMERS = ["华鑫铝业", "永达五金", "顺发配件", "金龙机械", "德盛电子"]

# ============================================================================
# 样式
# ============================================================================

class S:
    """样式常量"""
    TF = Font(name='微软雅黑', size=16, bold=True, color='4472C4')
    TF2 = Font(name='微软雅黑', size=16, bold=True, color='C00000')
    HF = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    HF2 = PatternFill(start_color='C00000', end_color='C00000', fill_type='solid')
    HFONT = Font(name='微软雅黑', size=10, bold=True, color='FFFFFF')
    BOLD = Font(name='微软雅黑', size=10, bold=True, color='0070C0')
    RED = Font(name='微软雅黑', size=10, bold=True, color='C00000')
    GRAY = Font(name='微软雅黑', size=9, color='808080')
    NORMAL = Font(name='微软雅黑', size=10)
    B = Border(left=Side('thin'), right=Side('thin'), top=Side('thin'), bottom=Side('thin'))
    RED_BG = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
    YELLOW_BG = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
    GREEN_BG = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')

# ============================================================================
# 终端色彩系统
# ============================================================================

class C:
    """ANSI终端颜色（跨平台）"""
    _enabled = None
    @classmethod
    def enabled(cls):
        if cls._enabled is None:
            cls._enabled = hasattr(sys.stdout, 'isatty') and sys.stdout.isatty()
            if os.name == 'nt':
                try:
                    import ctypes
                    ctypes.windll.kernel32.SetConsoleMode(
                        ctypes.windll.kernel32.GetStdHandle(-11), 7)
                    cls._enabled = True
                except Exception:
                    cls._enabled = False
        return cls._enabled
    # 前景色
    R = '\033[91m'     # 红
    G = '\033[92m'     # 绿
    Y = '\033[93m'     # 黄
    B = '\033[94m'     # 蓝
    M = '\033[95m'     # 紫
    C = '\033[96m'     # 青
    W = '\033[97m'     # 白
    D = '\033[90m'     # 灰
    # 样式
    BD = '\033[1m'     # 粗体
    DM = '\033[2m'     # 暗淡
    RST = '\033[0m'    # 重置

def cr(text: str, color: str) -> str:
    """给文本上色（仅在终端模式下生效）"""
    if C.enabled():
        return f"{color}{text}{C.RST}"
    return text

# 预定义彩色文本
def c_title(t):   return cr(t, C.BD + C.B)
def c_ok(t):      return cr(t, C.G)
def c_err(t):     return cr(t, C.R)
def c_warn(t):    return cr(t, C.Y)
def c_info(t):    return cr(t, C.C)
def c_dim(t):     return cr(t, C.D)
def c_head(t):    return cr(t, C.BD + C.M)
def c_num(t):     return cr(t, C.BD + C.G)
def c_money(t):   return cr(t, C.BD + C.Y)

# ============================================================================
# 界面美化工具
# ============================================================================

def print_title(text: str, width: int = 48):
    """打印功能标题框"""
    inner = f"  {text}  "
    pad = width - len(inner) - 2
    left = pad // 2
    right = pad - left
    print()
    print(cr(f"┌{'─' * left}{inner}{'─' * right}┐", C.B))
    print(cr(f"│{inner:^{width}}│", C.B))

def print_subtitle(text: str, width: int = 48):
    """打印副标题行"""
    print(cr(f"└{'─' * width}┘", C.B))

def print_table(headers: list, rows: list, col_widths: list = None):
    """打印对齐表格"""
    if not rows and not headers:
        return
    # 自动计算列宽
    if not col_widths:
        col_widths = []
        for i, h in enumerate(headers):
            max_w = len(str(h))
            for row in rows:
                if i < len(row):
                    max_w = max(max_w, len(str(row[i])))
            col_widths.append(min(max_w + 2, 30))
    # 表头
    header_line = ""
    for i, h in enumerate(headers):
        w = col_widths[i] if i < len(col_widths) else 12
        header_line += cr(f" {str(h):<{w}}", C.BD + C.W)
    sep = "─" * sum(min(w, 30) + 1 for w in col_widths)
    print(f"  {cr(sep, C.D)}")
    print(f"  {header_line}")
    print(f"  {cr(sep, C.D)}")
    # 数据行
    for ri, row in enumerate(rows):
        line = ""
        for i, val in enumerate(row):
            w = col_widths[i] if i < len(col_widths) else 12
            s = str(val) if val is not None else ""
            line += f" {s:<{w}}"
        # 交替行色
        if ri % 2 == 1:
            print(f"  {c_dim(line)}")
        else:
            print(f"  {line}")
    print(f"  {cr(sep, C.D)}")

def print_kv(pairs: list, indent: int = 4):
    """打印键值对列表 [(key, value), ...]"""
    pad = " " * indent
    max_key = max(len(str(k)) for k, v in pairs) if pairs else 0
    for k, v in pairs:
        print(f"{pad}{cr(str(k), C.W)}{' ' * (max_key - len(str(k)) + 2)}{v}")

def print_bar(label: str, value: float, max_val: float, width: int = 20):
    """打印进度条"""
    pct = value / max_val if max_val > 0 else 0
    filled = int(pct * width)
    bar = cr("█" * filled, C.G) + cr("░" * (width - filled), C.D)
    print(f"  {label:10s} {bar} {pct*100:5.1f}%")

def print_divider(char: str = "─", width: int = 48):
    """打印分隔线"""
    print(cr(char * width, C.D))

# ============================================================================
# 工具函数
# ============================================================================

def input_num(prompt: str, default: float = 0) -> float:
    """安全输入数字"""
    while True:
        val = input(prompt).strip()
        if val == "":
            return default
        try:
            return float(val)
        except ValueError:
            print("  请输入有效数字！")

def input_date(prompt: str, default: str = None) -> str:
    """安全输入日期"""
    if default is None:
        default = date.today().strftime("%Y-%m-%d")
    val = input(f"{prompt}（默认{default}）: ").strip()
    return val if val else default

def input_choice(prompt: str, options: List[str], default: str = "") -> str:
    """从选项中选择"""
    print(f"  可选：{', '.join(options)}")
    while True:
        val = input(f"  {prompt}（默认{default}）: ").strip()
        if val == "":
            return default
        if val in options:
            return val
        # 模糊匹配
        for opt in options:
            if val in opt:
                return opt
        print(f"  无效选项，请从 {', '.join(options)} 中选择")

def fmt(amount: float) -> str:
    """格式化金额"""
    return f"{amount:,.2f}"

def pause(msg: str = "\n按回车键继续..."):
    try:
        input(msg)
    except EOFError:
        pass

def clear():
    os.system('cls' if os.name == 'nt' else 'clear')

# ============================================================================
# 核心数据管理
# ============================================================================

class DataManager:
    """Excel数据管理"""

    def __init__(self):
        os.makedirs(DATA_DIR, exist_ok=True)
        os.makedirs(BACKUP_DIR, exist_ok=True)
        self.wb = None
        self._load()

    def _load(self):
        if os.path.exists(EXCEL_FILE):
            try:
                self.wb = load_workbook(EXCEL_FILE)
            except Exception:
                # 文件损坏时备份原文件
                bak = EXCEL_FILE + ".corrupted"
                if os.path.exists(bak):
                    os.remove(bak)
                os.rename(EXCEL_FILE, bak)
                print(f"  ⚠ 数据文件损坏，已备份为 {bak}")
                self.wb = Workbook()
                self.wb.remove(self.wb.active)
        else:
            self.wb = Workbook()
            self.wb.remove(self.wb.active)

    def save(self):
        if self.wb:
            try:
                self.wb.save(EXCEL_FILE)
            except Exception as e:
                print(f"  ✗ 保存失败：{e}")

    def sheet(self, name: str, create: bool = True):
        if name in self.wb.sheetnames:
            return self.wb[name]
        if create:
            return self.wb.create_sheet(name)
        return None

    def last_row(self, ws, col: int = 1) -> int:
        if not ws:
            return 0
        r = ws.max_row
        while r > 1 and ws.cell(row=r, column=col).value is None:
            r -= 1
        return r

    def next_row(self, ws, col: int = 1) -> int:
        return self.last_row(ws, col) + 1

    def read_rows(self, ws, start_row: int = 4, end_col: int = 20) -> List[list]:
        """读取所有数据行"""
        rows = []
        if not ws:
            return rows
        for r in range(start_row, self.last_row(ws) + 1):
            row = []
            for c in range(1, end_col + 1):
                row.append(ws.cell(row=r, column=c).value)
            rows.append(row)
        return rows

    def write_row(self, ws, row: int, data: list, start_col: int = 1):
        """写入一行数据"""
        for i, val in enumerate(data):
            ws.cell(row=row, column=start_col + i, value=val)
            ws.cell(row=row, column=start_col + i).border = S.B

    def set_header(self, ws, row: int, headers: list, fill=None):
        """设置表头"""
        for col, h in enumerate(headers, 1):
            c = ws.cell(row=row, column=col, value=h)
            c.fill = fill or S.HF
            c.font = S.HFONT
            c.border = S.B
            c.alignment = Alignment(horizontal='center', vertical='center')

    def set_title(self, ws, title: str, subtitle: str = "", font=None, merge_end: str = None):
        """设置标题"""
        ws['A1'] = title
        ws['A1'].font = font or S.TF
        end = merge_end or get_column_letter(len(self._get_headers(ws)))
        ws.merge_cells(f'A1:{end}1')
        if subtitle:
            ws['A2'] = subtitle
            ws['A2'].font = S.GRAY
            ws.merge_cells(f'A2:{end}2')

    def _get_headers(self, ws):
        """获取表头"""
        headers = []
        if ws:
            c = 1
            while ws.cell(row=3, column=c).value:
                headers.append(ws.cell(row=3, column=c).value)
                c += 1
        return headers

# ============================================================================
# 工作表初始化
# ============================================================================

def init_all_sheets(dm: DataManager):
    """初始化所有工作表"""

    # 1. 首页
    ws = dm.sheet("首页")
    ws.sheet_properties.tabColor = "4472C4"
    ws['B2'] = "小型氧化加工厂财务管理系统"
    ws['B2'].font = S.TF
    ws.merge_cells('B2:F2')
    ws['B3'] = VERSION
    ws['B3'].font = S.GRAY
    ws.merge_cells('B3:F3')
    ws['B5'] = "◆ 快速开始"
    ws['B5'].font = S.RED
    for i, t in enumerate(["1. 双击 启动财务系统.bat 或 python 氧化加工厂财务系统.py",
                            "2. 首次运行自动初始化，生成模拟数据可选",
                            "3. 按菜单编号选择功能，数据自动保存"], 6):
        ws[f'B{i}'] = t
        ws[f'B{i}'].font = S.NORMAL
    ws['B10'] = "◆ 功能总览（45项）"
    ws['B10'].font = S.RED
    cats = [("日常记账","1-5"),("报表生成","6-10"),("税务管理","11-16"),
            ("工资社保","17-19"),("资产管理","20-22"),("成本分析","23-26"),
            ("预算管理","27-29"),("数据工具","30-33"),("智能提醒","34-36"),
            ("常用模板","37-41"),("系统功能","42-45")]
    for i, (n, nums) in enumerate(cats, 11):
        ws[f'B{i}'] = n
        ws[f'B{i}'].font = S.BOLD
        ws[f'C{i}'] = f"[{nums}]"
        ws[f'C{i}'].font = S.GRAY
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 10

    # 2. 收入记录
    ws = dm.sheet("收入记录")
    ws.sheet_properties.tabColor = "00B050"
    dm.set_title(ws, "收入记录（对账后金额）", "含税金额=不含税金额+税额", merge_end="M")
    dm.set_header(ws, 3, ["日期","凭证号","客户名称","收入金额","收款方式",
        "含税/不含税","税率","税额","不含税金额","开票状态","对冲/代付类型","实际付款方","备注"])
    for i, w in enumerate([12,10,15,14,12,12,8,12,14,10,14,15,20], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    try:
        dv = DataValidation(type="list", formula1='"含税,不含税"')
        dv.add('F4:F2000'); ws.add_data_validation(dv)
        dv2 = DataValidation(type="list", formula1='"已开票,未开票,部分开票"')
        dv2.add('J4:J2000'); ws.add_data_validation(dv2)
        dv3 = DataValidation(type="list", formula1='"正常,对冲货款,代付货款"')
        dv3.add('K4:K2000'); ws.add_data_validation(dv3)
        dv4 = DataValidation(type="list", formula1='"银行转账,微信,支付宝,现金"')
        dv4.add('E4:E2000'); ws.add_data_validation(dv4)
    except: pass

    # 3. 支出记录
    ws = dm.sheet("支出记录")
    ws.sheet_properties.tabColor = "FF0000"
    dm.set_title(ws, "支出记录", "", merge_end="G")
    dm.set_header(ws, 3, ["日期","凭证号","类别","金额","供应商","付款方式","备注"])
    for i, w in enumerate([12,10,15,12,18,12,25], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    try:
        dv = DataValidation(type="list", formula1=f'"{",".join(EXPENSE_CATEGORIES)}"')
        dv.add('C4:C2000'); ws.add_data_validation(dv)
    except Exception:
        pass

    # 4. 应收应付
    ws = dm.sheet("应收应付")
    ws.sheet_properties.tabColor = "7030A0"
    dm.set_title(ws, "应收应付", "", merge_end="G")
    dm.set_header(ws, 3, ["客户名称","期初应收","本期增加","本期收款","对冲减少","期末应收","备注"])

    # 5. 利润分析表
    ws = dm.sheet("利润分析表")
    ws.sheet_properties.tabColor = "C00000"
    dm.set_title(ws, "利润分析表", "", font=S.TF2, merge_end="E")
    dm.set_header(ws, 3, ["序号","科目","行次","金额(元)","备注"], fill=S.HF2)
    items = [(1,"收入：月总产量",1,0),(2,"厂租",1,0),(3,"电费",2,0),(4,"水费",3,0),
             (5,"化工",4,0),(6,"染料封闭剂",5,0),(7,"工资",6,0),(8,"包装材料",7,0),
             (9,"货车运输费",8,0),(10,"外发加工费",9,0),(11,"税金",10,0),(12,"社保费",11,0),
             (13,"业务费",12,0),(14,"办公用品费",13,0),(15,"维修费",14,0),(16,"挂具费",15,0),
             ("","支出合计","",None),("","净利润","",None)]
    for i, (no, item, rn, amt) in enumerate(items, 4):
        ws.cell(row=i, column=1, value=no).border = S.B
        ws.cell(row=i, column=2, value=item).border = S.B
        ws.cell(row=i, column=3, value=rn).border = S.B
        if amt is not None:
            ws.cell(row=i, column=4, value=amt).border = S.B
            ws.cell(row=i, column=4).number_format = '#,##0.00'
    ws.cell(row=20, column=2).font = S.RED
    ws.cell(row=21, column=2).font = Font(name='微软雅黑', size=11, bold=True, color='C00000')
    ws.cell(row=21, column=4, value="=D4-D20")
    ws.cell(row=21, column=4).font = Font(name='微软雅黑', size=11, bold=True, color='C00000')
    ws.cell(row=21, column=4).number_format = '#,##0.00'

    # 6. 工资表
    ws = dm.sheet("工资表")
    ws.sheet_properties.tabColor = "FFC000"
    dm.set_title(ws, "工资表", f"年月：{date.today().strftime('%Y年%m月')}", merge_end="K")
    dm.set_header(ws, 4, ["序号","姓名","基本工资","加班费","奖金","应发合计",
        "社保个人","公积金个人","个税","实发合计","发放方式"])

    # 7. 发票登记表
    ws = dm.sheet("发票登记表")
    ws.sheet_properties.tabColor = "00B0F0"
    dm.set_title(ws, "发票登记表", "", merge_end="I")
    dm.set_header(ws, 3, ["开票日期","发票号码","发票类型","客户名称","金额","税额","价税合计","发票状态","备注"])

    # 8. 银行对账表
    ws = dm.sheet("银行对账表")
    ws.sheet_properties.tabColor = "92D050"
    dm.set_title(ws, "银行对账表", "", merge_end="J")
    dm.set_header(ws, 4, ["日期","摘要","银行收入","银行支出","银行余额",
        "账面收入","账面支出","账面余额","差异","备注"])

    # 9. 材料进销存台账
    ws = dm.sheet("材料进销存台账")
    ws.sheet_properties.tabColor = "FF6699"
    dm.set_title(ws, "材料进销存台账", f"常用材料：{'/'.join(MATERIALS)}", merge_end="L")
    dm.set_header(ws, 3, ["日期","材料名称","规格","入库数量","入库单价","入库金额",
        "出库数量","出库单价","出库金额","结存数量","结存金额","备注"])

    # 10. 月度经营报表
    ws = dm.sheet("月度经营报表")
    ws.sheet_properties.tabColor = "C00000"
    dm.set_title(ws, "月度经营报表", "", font=S.TF2, merge_end="D")
    dm.set_header(ws, 4, ["项目","本月金额","上月金额","环比增减"])

    # 11. 使用帮助
    ws = dm.sheet("使用帮助")
    ws.sheet_properties.tabColor = "808080"
    ws['A1'] = f"使用帮助 - {VERSION}"
    ws['A1'].font = S.TF2
    help_lines = [
        "","◆ 运行方法","  python 氧化加工厂财务系统.py","  或双击 启动财务系统.bat","",
        "◆ 依赖安装","  pip install openpyxl","",
        "◆ 45个功能一览","",
        "【日常记账 1-5】","  1.录入收入 2.录入支出 3.收入汇总 4.支出汇总 5.批量导入","",
        "【报表生成 6-10】","  6.利润报表 7.资产负债表 8.月度报表 9.客户对账单 10.多期对比","",
        "【税务管理 11-16】","  11.增值税 12.所得税 13.六税两费 14.残保金 15.工会经费 16.税务汇总","",
        "【工资社保 17-19】","  17.工资计算 18.个税计算 19.社保明细","",
        "【资产管理 20-22】","  20.固定资产 21.折旧计算 22.低值易耗品","",
        "【成本分析 23-26】","  23.成本核算 24.差异分析 25.应收账龄 26.应付账龄","",
        "【预算管理 27-29】","  27.预算编制 28.预算控制 29.预算预警","",
        "【数据工具 30-33】","  30.多条件查询 31.导入导出 32.自动备份 33.数据修复","",
        "【智能提醒 34-36】","  34.应收逾期 35.库存预警 36.利润预警","",
        "【常用模板 37-41】","  37.工资表模板 38.发票登记 39.银行对账 40.材料台账 41.月度报表","",
        "【系统 42-45】","  42.模拟数据 43.系统设置 44.使用帮助 45.关于系统","",
        "◆ 数据存储","  数据文件：财务数据/财务数据.xlsx","  备份目录：备份/",
    ]
    for i, line in enumerate(help_lines, 3):
        ws.cell(row=i, column=1, value=line)
        if line.startswith("◆"):
            ws.cell(row=i, column=1).font = S.RED
        elif line.startswith("【"):
            ws.cell(row=i, column=1).font = S.BOLD
    ws.column_dimensions['A'].width = 55

    dm.save()
    print(f"  {c_ok('所有工作表初始化完成（11个）')}")

# ============================================================================
# 功能1-5：日常记账
# ============================================================================

def func_1_add_income(dm: DataManager):
    """录入收入"""
    print_title("录入收入")
    dt = input_date("  日期")
    customer = input("  客户名称: ").strip()
    if not customer:
        print("  " + c_err('客户名称不能为空')); return
    amount = input_num("  金额: ")
    if amount <= 0:
        print("  " + c_err('金额必须大于0')); return
    pay = input_choice("  收款方式", ["银行转账","微信","支付宝","现金"], "银行转账")
    tax_type = input_choice("  含税/不含税", ["含税","不含税"], "含税")
    invoice = input_choice("  开票状态", ["已开票","未开票","部分开票"], "未开票")
    remark = input("  备注: ").strip()

    ws = dm.sheet("收入记录")
    row = dm.next_row(ws)
    voucher = f"收-{datetime.now().strftime('%m')}-{row-3:03d}"
    rate = 0.03
    if tax_type == "含税":
        tax_amt = round(amount * rate / (1 + rate), 2)
        no_tax = round(amount / (1 + rate), 2)
    else:
        tax_amt = round(amount * rate, 2)
        no_tax = amount

    dm.write_row(ws, row, [dt, voucher, customer, amount, pay, tax_type, rate,
                           tax_amt, no_tax, invoice, "正常", "", remark])
    dm.save()
    print(f"  ✓ 收入已录入，凭证号：{voucher}，税额：{fmt(tax_amt)}")

def func_2_add_expense(dm: DataManager):
    """录入支出"""
    print_title("录入支出")
    dt = input_date("  日期")
    cat = input_choice("  类别", EXPENSE_CATEGORIES)
    if not cat:
        print("  " + c_err('类别不能为空')); return
    amount = input_num("  金额: ")
    if amount <= 0:
        print("  " + c_err('金额必须大于0')); return
    supplier = input("  供应商: ").strip()
    pay = input_choice("  付款方式", ["银行转账","微信","支付宝","现金"], "银行转账")
    remark = input("  备注: ").strip()

    ws = dm.sheet("支出记录")
    row = dm.next_row(ws)
    voucher = f"支-{datetime.now().strftime('%m')}-{row-3:03d}"
    dm.write_row(ws, row, [dt, voucher, cat, amount, supplier, pay, remark])
    dm.save()
    print(f"  ✓ 支出已录入，凭证号：{voucher}")

def func_3_income_summary(dm: DataManager):
    """收入汇总"""
    print_title("收入汇总")
    ws = dm.sheet("收入记录", create=False)
    if not ws or dm.last_row(ws) < 4:
        print(f"\n  {c_warn('暂无收入数据')}"); return

    total = 0; count = 0; by_cust = {}
    for r in range(4, dm.last_row(ws) + 1):
        amt = ws.cell(row=r, column=4).value
        cust = ws.cell(row=r, column=3).value
        if amt and isinstance(amt, (int, float)):
            total += amt; count += 1
            by_cust[cust] = by_cust.get(cust, 0) + amt

    print(f"\n  {c_info('总收入')}: {c_money(fmt(total))} 元")
    print(f"  {c_info('记录数')}: {c_num(str(count))} 条")
    if by_cust:
        print(f"\n  {c_head('按客户汇总')}:")
        rows = []
        for c_name, a in sorted(by_cust.items(), key=lambda x: -x[1]):
            pct = a / total * 100 if total else 0
            rows.append([c_name, f"{fmt(a)} 元", f"{pct:.1f}%"])
        print_table(["客户名称", "收入金额", "占比"], rows)

def func_4_expense_summary(dm: DataManager):
    """支出汇总"""
    print_title("支出汇总")
    ws = dm.sheet("支出记录", create=False)
    if not ws or dm.last_row(ws) < 4:
        print(f"\n  {c_warn('暂无支出数据')}"); return

    total = 0; count = 0; by_cat = {}
    for r in range(4, dm.last_row(ws) + 1):
        amt = ws.cell(row=r, column=4).value
        cat = ws.cell(row=r, column=3).value
        if amt and isinstance(amt, (int, float)):
            total += amt; count += 1
            by_cat[cat] = by_cat.get(cat, 0) + amt

    print(f"\n  {c_info('总支出')}: {c_money(fmt(total))} 元")
    print(f"  {c_info('记录数')}: {c_num(str(count))} 条")
    if by_cat:
        print(f"\n  {c_head('按类别汇总')}:")
        rows = []
        for cat_name, a in sorted(by_cat.items(), key=lambda x: -x[1]):
            pct = a / total * 100 if total else 0
            rows.append([cat_name, f"{fmt(a)} 元", f"{pct:.1f}%"])
        print_table(["支出类别", "支出金额", "占比"], rows)

def func_5_batch_import(dm: DataManager):
    """批量导入"""
    print_title("批量导入")
    fp = input("  CSV文件路径: ").strip()
    if not fp or not os.path.exists(fp):
        print("  " + c_err('文件不存在')); return
    target = input_choice("  导入到", ["收入记录","支出记录"])
    if not target:
        return
    ws = dm.sheet(target, create=False)
    if not ws:
        print(f"  ✗ 未找到{target}表"); return

    try:
        with open(fp, 'r', encoding='utf-8-sig') as f:
            reader = csv.reader(f)
            next(reader)  # 跳过标题
            count = 0
            for row_data in reader:
                r = dm.next_row(ws)
                dm.write_row(ws, r, row_data)
                count += 1
        dm.save()
        print(f"  ✓ 导入完成，共 {count} 条记录")
    except Exception as e:
        print(f"  ✗ 导入失败：{e}")

# ============================================================================
# 功能6-10：报表生成
# ============================================================================

def func_6_profit_report(dm: DataManager):
    """利润报表"""
    print_title("利润报表")

    # 收入
    ws_inc = dm.sheet("收入记录", create=False)
    income = 0
    if ws_inc:
        for r in range(4, dm.last_row(ws_inc) + 1):
            v = ws_inc.cell(row=r, column=4).value
            if v and isinstance(v, (int, float)):
                income += v

    # 支出
    ws_exp = dm.sheet("支出记录", create=False)
    expense = 0; by_cat = {}
    if ws_exp:
        for r in range(4, dm.last_row(ws_exp) + 1):
            v = ws_exp.cell(row=r, column=4).value
            c = ws_exp.cell(row=r, column=3).value
            if v and isinstance(v, (int, float)):
                expense += v
                by_cat[c] = by_cat.get(c, 0) + v

    profit = income - expense
    rate = profit / income * 100 if income > 0 else 0

    print(f"\n  收入合计：{fmt(income)} 元")
    print(f"  支出合计：{fmt(expense)} 元")
    print(f"  ─────────────────────")
    print(f"  净利润：  {fmt(profit)} 元")
    print(f"  利润率：  {rate:.1f}%")

    if by_cat:
        print("\n  支出明细：")
        for c, a in sorted(by_cat.items(), key=lambda x: -x[1]):
            pct = a / expense * 100 if expense else 0
            bar = "█" * int(pct / 5)
            print(f"    {c:10s} {fmt(a):>12s} ({pct:5.1f}%) {bar}")

    if profit < 0:
        print("\n  ⚠ 警告：当前亏损！")
    elif rate > 30:
        print("\n  ✓ 利润率良好！")
    elif rate < 10:
        print("\n  ⚠ 利润率偏低")

def func_7_balance_sheet(dm: DataManager):
    """资产负债表"""
    print_title("资产负债表（简化版）")

    ws_ar = dm.sheet("应收应付", create=False)
    total_ar = 0
    if ws_ar:
        for r in range(4, dm.last_row(ws_ar) + 1):
            v = ws_ar.cell(row=r, column=6).value
            if v and isinstance(v, (int, float)):
                total_ar += v

    print(f"\n  资产：")
    print(f"    应收账款：{fmt(total_ar)} 元")
    print(f"    资产合计：{fmt(total_ar)} 元")
    print(f"\n  所有者权益：")
    print(f"    未分配利润：{fmt(total_ar)} 元")

def func_8_monthly_report(dm: DataManager):
    """月度报表"""
    print_title("月度经营报表")
    month = input("  输入月份（YYYY-MM，默认本月）: ").strip() or datetime.now().strftime("%Y-%m")

    ws_inc = dm.sheet("收入记录", create=False)
    ws_exp = dm.sheet("支出记录", create=False)

    inc_total = 0; exp_total = 0
    if ws_inc:
        for r in range(4, dm.last_row(ws_inc) + 1):
            dt = ws_inc.cell(row=r, column=1).value
            amt = ws_inc.cell(row=r, column=4).value
            if dt and amt and isinstance(amt, (int, float)):
                if str(dt).startswith(month):
                    inc_total += amt
    if ws_exp:
        for r in range(4, dm.last_row(ws_exp) + 1):
            dt = ws_exp.cell(row=r, column=1).value
            amt = ws_exp.cell(row=r, column=4).value
            if dt and amt and isinstance(amt, (int, float)):
                if str(dt).startswith(month):
                    exp_total += amt

    profit = inc_total - exp_total
    print(f"\n  月份：{month}")
    print(f"  收入：{fmt(inc_total)} 元")
    print(f"  支出：{fmt(exp_total)} 元")
    print(f"  利润：{fmt(profit)} 元")

def func_9_customer_statement(dm: DataManager):
    """客户对账单"""
    print_title("客户对账单")
    customer = input("  客户名称: ").strip()
    if not customer:
        return

    ws = dm.sheet("收入记录", create=False)
    if not ws:
        print("  暂无数据"); return

    total = 0; count = 0
    print(f"\n  {'日期':<12} {'凭证号':<10} {'金额':>12} {'状态':<8}")
    print("  " + "─" * 45)
    for r in range(4, dm.last_row(ws) + 1):
        cust = str(ws.cell(row=r, column=3).value or "")
        if customer in cust:
            dt = ws.cell(row=r, column=1).value
            vno = ws.cell(row=r, column=2).value
            amt = ws.cell(row=r, column=4).value
            status = ws.cell(row=r, column=10).value
            if amt and isinstance(amt, (int, float)):
                total += amt; count += 1
                print(f"  {str(dt):<12} {str(vno):<10} {fmt(amt):>12} {str(status or ''):<8}")

    print("  " + "─" * 45)
    print(f"  合计：{fmt(total)} 元（{count}笔）")

def func_10_multi_period(dm: DataManager):
    """多期对比"""
    print_title("多期对比分析")
    months = input("  输入月份（逗号分隔，如 2024-03,2024-04,2024-05）: ").strip()
    if not months:
        return
    month_list = [m.strip() for m in months.split(",")]

    ws_inc = dm.sheet("收入记录", create=False)
    ws_exp = dm.sheet("支出记录", create=False)

    print(f"\n  {'月份':<12} {'收入':>12} {'支出':>12} {'利润':>12} {'利润率':>8}")
    print("  " + "─" * 60)
    for month in month_list:
        inc = exp = 0
        if ws_inc:
            for r in range(4, dm.last_row(ws_inc) + 1):
                dt = ws_inc.cell(row=r, column=1).value
                amt = ws_inc.cell(row=r, column=4).value
                if dt and amt and isinstance(amt, (int, float)) and str(dt).startswith(month):
                    inc += amt
        if ws_exp:
            for r in range(4, dm.last_row(ws_exp) + 1):
                dt = ws_exp.cell(row=r, column=1).value
                amt = ws_exp.cell(row=r, column=4).value
                if dt and amt and isinstance(amt, (int, float)) and str(dt).startswith(month):
                    exp += amt
        profit = inc - exp
        rate = profit / inc * 100 if inc > 0 else 0
        print(f"  {month:<12} {fmt(inc):>12} {fmt(exp):>12} {fmt(profit):>12} {rate:>7.1f}%")

# ============================================================================
# 功能11-16：税务管理
# ============================================================================

def func_11_vat(dm: DataManager):
    """增值税计算"""
    print_title("增值税计算（小规模纳税人）")
    sales = input_num("  月销售额: ")
    if sales <= 100000:
        print(f"\n  ✓ 月销售额≤10万元，免征增值税！")
        print(f"  应缴税额：0.00 元")
    else:
        tax = sales * 0.03
        print(f"\n  应税销售额：{fmt(sales)} 元")
        print(f"  税率：3%")
        print(f"  应缴增值税：{fmt(tax)} 元")

def func_12_income_tax(dm: DataManager):
    """所得税计算"""
    print_title("企业所得税计算（小微企业优惠）")
    profit = input_num("  年应纳税所得额: ")
    if profit <= 1000000:
        tax = profit * 0.025
        note = "≤100万，税率2.5%"
    elif profit <= 3000000:
        tax = 25000 + (profit - 1000000) * 0.05
        note = "100-300万，税率5%"
    else:
        tax = profit * 0.25
        note = ">300万，税率25%"
    print(f"\n  应纳税所得额：{fmt(profit)} 元")
    print(f"  应缴所得税：{fmt(tax)} 元")
    rate_str = f"{tax/profit*100:.2f}%" if profit > 0 else "N/A"
    print(f"  实际税率：{rate_str}")
    print(f"  政策：{note}")

def func_13_six_taxes(dm: DataManager):
    """六税两费减免"""
    print_title("六税两费减免计算")
    print("  小规模纳税人可减半征收：")
    print("    资源税、城市维护建设税、房产税")
    print("    城镇土地使用税、印花税、教育费附加")
    base = input_num("  计税基数: ")
    half = base * 0.5
    print(f"\n  原应缴：{fmt(base)} 元")
    print(f"  减免后：{fmt(half)} 元")
    print(f"  节省：{fmt(base - half)} 元")

def func_14_disabled_fund(dm: DataManager):
    """残保金计算"""
    print_title("残保金计算")
    employees = int(input_num("  在职职工人数: "))
    avg_salary = input_num("  月平均工资: ")
    disabled = int(input_num("  残疾职工人数: "))
    ratio = disabled / employees if employees > 0 else 0

    if employees <= 30:
        print(f"\n  ✓ 职工人数≤30人，免征残保金！")
    else:
        required = employees * 0.015
        gap = max(0, required - disabled)
        annual_fund = gap * avg_salary * 12
        print(f"\n  安置比例：{ratio*100:.1f}%（要求1.5%）")
        print(f"  应安置：{required:.1f}人，实际：{disabled}人")
        print(f"  差额：{gap:.1f}人")
        print(f"  年应缴残保金：{fmt(annual_fund)} 元")

def func_15_union_fund(dm: DataManager):
    """工会经费计算"""
    print_title("工会经费计算")
    total_salary = input_num("  工资总额: ")
    rate = 0.02
    amount = total_salary * rate
    print(f"\n  工资总额：{fmt(total_salary)} 元")
    print(f"  工会经费（2%）：{fmt(amount)} 元")

def func_16_tax_summary(dm: DataManager):
    """税务汇总"""
    print_title("税务综合汇总")

    # 从支出中统计税金
    ws = dm.sheet("支出记录", create=False)
    tax_total = 0; si_total = 0
    if ws:
        for r in range(4, dm.last_row(ws) + 1):
            cat = ws.cell(row=r, column=3).value
            amt = ws.cell(row=r, column=4).value
            if amt and isinstance(amt, (int, float)):
                if cat == "税金":
                    tax_total += amt
                elif cat == "社保费":
                    si_total += amt

    print(f"\n  已缴税金：{fmt(tax_total)} 元")
    print(f"  已缴社保：{fmt(si_total)} 元")
    print(f"  税费合计：{fmt(tax_total + si_total)} 元")

# ============================================================================
# 功能17-19：工资社保
# ============================================================================

def func_17_payroll(dm: DataManager):
    """工资计算"""
    print_title("工资计算")
    base = input_num("  基本工资: ")
    overtime = input_num("  加班费（默认0）: ", 0)
    bonus = input_num("  奖金（默认0）: ", 0)
    si_rate = input_num("  社保个人比例（默认10%）: ", 0.1)

    gross = base + overtime + bonus
    si = gross * si_rate
    taxable = gross - si - 5000
    tax = calc_income_tax(taxable)
    net = gross - si - tax

    print(f"\n  应发合计：{fmt(gross)} 元")
    print(f"  社保个人：-{fmt(si)} 元")
    print(f"  个人所得税：-{fmt(tax)} 元")
    print(f"  ─────────────────")
    print(f"  实发合计：{fmt(net)} 元")

    # 写入工资表
    ws = dm.sheet("工资表")
    row = dm.next_row(ws, 2)
    dm.write_row(ws, row, [row - 4, "", base, overtime, bonus, gross, si, 0, tax, net, "银行转账"])
    dm.save()
    print(f"  ✓ 已写入工资表第{row}行")

def calc_income_tax(taxable: float) -> float:
    """计算个人所得税（七级超额累进税率）"""
    if taxable <= 0:
        return 0
    brackets = [
        (36000, 0.03, 0),
        (144000, 0.10, 2520),
        (300000, 0.20, 16920),
        (420000, 0.25, 31920),
        (660000, 0.30, 52920),
        (960000, 0.35, 85920),
        (float('inf'), 0.45, 181920),
    ]
    for limit, rate, deduction in brackets:
        if taxable <= limit:
            return round(taxable * rate - deduction, 2)
    return 0

def func_18_tax_calc(dm: DataManager):
    """个税计算"""
    print_title("个人所得税计算")
    income = input_num("  税前收入: ")
    si = input_num("  社保扣除（默认0）: ", 0)
    threshold = 5000
    taxable = income - si - threshold
    if taxable <= 0:
        print(f"\n  应纳税所得额：{fmt(taxable)} 元")
        print(f"  ✓ 无需缴纳个税")
    else:
        tax = calc_income_tax(taxable)
        print(f"\n  税前收入：{fmt(income)} 元")
        print(f"  社保扣除：-{fmt(si)} 元")
        print(f"  起征点：  -{fmt(threshold)} 元")
        print(f"  应纳税额：{fmt(taxable)} 元")
        print(f"  个人所得税：{fmt(tax)} 元")
        print(f"  税后收入：{fmt(income - si - tax)} 元")

def func_19_social_insurance(dm: DataManager):
    """社保明细"""
    print_title("社保明细")
    base = input_num("  缴费基数: ")
    print(f"\n  缴费基数：{fmt(base)} 元")
    print(f"  {'项目':<10} {'个人比例':>8} {'个人金额':>10} {'单位比例':>8} {'单位金额':>10}")
    print("  " + "─" * 50)
    items = [("养老保险","8%","16%"),("医疗保险","2%","8%"),("失业保险","0.5%","0.5%"),("工伤保险","0%","0.2%"),("生育保险","0%","0.8%")]
    total_p = 0; total_c = 0
    for name, pr, cr in items:
        p_amt = base * float(pr.strip('%')) / 100
        c_amt = base * float(cr.strip('%')) / 100
        total_p += p_amt; total_c += c_amt
        print(f"  {name:<10} {pr:>8} {fmt(p_amt):>10} {cr:>8} {fmt(c_amt):>10}")
    print("  " + "─" * 50)
    print(f"  {'合计':<10} {'10.5%':>8} {fmt(total_p):>10} {'25.5%':>8} {fmt(total_c):>10}")

# ============================================================================
# 功能20-22：资产管理
# ============================================================================

def func_20_fixed_assets(dm: DataManager):
    """固定资产登记"""
    print_title("固定资产登记")
    ws = dm.sheet("固定资产", create=True)
    if dm.last_row(ws) < 3:
        dm.set_title(ws, "固定资产台账", "", merge_end="H")
        dm.set_header(ws, 3, ["资产名称","购入日期","原值","预计残值率","使用年限","月折旧","累计折旧","净值"])

    name = input("  资产名称: ").strip()
    if not name: return
    dt = input_date("  购入日期")
    value = input_num("  原值: ")
    residual_rate = input_num("  预计残值率%（默认5）: ", 5) / 100
    years = int(input_num("  使用年限: "))

    monthly_dep = value * (1 - residual_rate) / (years * 12)
    row = dm.next_row(ws)
    dm.write_row(ws, row, [name, dt, value, residual_rate, years, round(monthly_dep, 2), 0, value])
    dm.save()
    print(f"  ✓ 已登记，月折旧：{fmt(monthly_dep)} 元")

def func_21_depreciation(dm: DataManager):
    """折旧计算"""
    print_title("折旧计算")
    value = input_num("  资产原值: ")
    residual_rate = input_num("  残值率%（默认5）: ", 5) / 100
    years = int(input_num("  使用年限: "))
    used_months = int(input_num("  已使用月数（默认0）: ", 0))

    monthly = value * (1 - residual_rate) / (years * 12)
    accumulated = monthly * used_months
    net = value - accumulated

    print(f"\n  资产原值：{fmt(value)} 元")
    print(f"  预计残值：{fmt(value * residual_rate)} 元")
    print(f"  月折旧额：{fmt(monthly)} 元")
    print(f"  累计折旧：{fmt(accumulated)} 元")
    print(f"  资产净值：{fmt(net)} 元")

def func_22_low_value(dm: DataManager):
    """低值易耗品管理"""
    print_title("低值易耗品管理（单价≤2000元）")
    ws = dm.sheet("低值易耗品", create=True)
    if dm.last_row(ws) < 3:
        dm.set_title(ws, "低值易耗品台账", "", merge_end="F")
        dm.set_header(ws, 3, ["物品名称","购入日期","单价","数量","金额","领用状态"])

    name = input("  物品名称: ").strip()
    if not name: return
    dt = input_date("  购入日期")
    price = input_num("  单价: ")
    qty = int(input_num("  数量: "))
    status = input_choice("  领用状态", ["在库","已领用","已报废"], "在库")

    row = dm.next_row(ws)
    dm.write_row(ws, row, [name, dt, price, qty, price * qty, status])
    dm.save()
    print(f"  ✓ 已登记，金额：{fmt(price * qty)} 元")

# ============================================================================
# 功能23-26：成本分析
# ============================================================================

def func_23_cost_accounting(dm: DataManager):
    """成本核算"""
    print_title("成本核算（15类支出归集）")
    ws = dm.sheet("支出记录", create=False)
    if not ws or dm.last_row(ws) < 4:
        print("  暂无支出数据"); return

    by_cat = {}
    for r in range(4, dm.last_row(ws) + 1):
        cat = ws.cell(row=r, column=3).value
        amt = ws.cell(row=r, column=4).value
        if cat and amt and isinstance(amt, (int, float)):
            by_cat[cat] = by_cat.get(cat, 0) + amt

    total = sum(by_cat.values())
    print(f"\n  总成本：{fmt(total)} 元")
    print(f"\n  {'类别':<12} {'金额':>12} {'占比':>8} {'趋势':>20}")
    print("  " + "─" * 55)
    for cat in EXPENSE_CATEGORIES:
        amt = by_cat.get(cat, 0)
        if amt > 0:
            pct = amt / total * 100 if total else 0
            bar = "█" * int(pct / 2)
            print(f"  {cat:<12} {fmt(amt):>12} {pct:>7.1f}% {bar}")

def func_24_variance_analysis(dm: DataManager):
    """差异分析"""
    print_title("成本差异分析")
    budget = input_num("  预算金额: ")
    ws = dm.sheet("支出记录", create=False)
    actual = 0
    if ws:
        for r in range(4, dm.last_row(ws) + 1):
            v = ws.cell(row=r, column=4).value
            if v and isinstance(v, (int, float)):
                actual += v

    diff = budget - actual
    pct = diff / budget * 100 if budget else 0
    print(f"\n  预算：{fmt(budget)} 元")
    print(f"  实际：{fmt(actual)} 元")
    print(f"  差异：{fmt(diff)} 元（{pct:+.1f}%）")
    if diff < 0:
        print(f"  ⚠ 超支 {fmt(abs(diff))} 元！")
    else:
        print(f"  ✓ 节余 {fmt(diff)} 元")

def func_25_ar_aging(dm: DataManager):
    """应收账龄分析"""
    print_title("应收账龄分析")
    ws = dm.sheet("应收应付", create=False)
    if not ws or dm.last_row(ws) < 4:
        print("  暂无数据"); return

    print(f"\n  {'客户':<12} {'期末应收':>12} {'状态':>10}")
    print("  " + "─" * 38)
    for r in range(4, dm.last_row(ws) + 1):
        cust = ws.cell(row=r, column=1).value
        amt = ws.cell(row=r, column=6).value
        if amt and isinstance(amt, (int, float)) and amt > 0:
            status = "⚠ 大额" if amt > 20000 else ("注意" if amt > 10000 else "正常")
            print(f"  {str(cust):<12} {fmt(amt):>12} {status:>10}")

def func_26_ap_aging(dm: DataManager):
    """应付账龄分析"""
    print_title("应付账龄分析")
    ws = dm.sheet("支出记录", create=False)
    if not ws or dm.last_row(ws) < 4:
        print("  暂无数据"); return

    by_supplier = {}
    for r in range(4, dm.last_row(ws) + 1):
        sup = ws.cell(row=r, column=5).value
        amt = ws.cell(row=r, column=4).value
        if sup and amt and isinstance(amt, (int, float)):
            by_supplier[sup] = by_supplier.get(sup, 0) + amt

    print(f"\n  {'供应商':<15} {'累计应付':>12}")
    print("  " + "─" * 30)
    for sup, amt in sorted(by_supplier.items(), key=lambda x: -x[1]):
        print(f"  {sup:<15} {fmt(amt):>12}")

# ============================================================================
# 功能27-29：预算管理
# ============================================================================

def func_27_budget(dm: DataManager):
    """预算编制"""
    print_title("预算编制")
    ws = dm.sheet("预算表", create=True)
    if dm.last_row(ws) < 3:
        dm.set_title(ws, "预算表", "", merge_end="D")
        dm.set_header(ws, 3, ["类别","月预算","年预算","备注"])

    print("  为各类支出设置月度预算：")
    for cat in EXPENSE_CATEGORIES:
        default_budget = {"厂租":8000,"电费":4000,"水费":400,"工资":15000,"社保费":2500}.get(cat, 1000)
        budget = input_num(f"  {cat}（默认{default_budget}）: ", default_budget)
        if budget > 0:
            row = dm.next_row(ws)
            dm.write_row(ws, row, [cat, budget, budget * 12, ""])
    dm.save()
    print(f"  {c_ok('预算编制完成')}")

def func_28_budget_control(dm: DataManager):
    """预算控制"""
    print_title("预算执行监控")
    ws_budget = dm.sheet("预算表", create=False)
    ws_exp = dm.sheet("支出记录", create=False)
    if not ws_budget or not ws_exp:
        print("  缺少预算表或支出数据"); return

    budgets = {}
    for r in range(4, dm.last_row(ws_budget) + 1):
        cat = ws_budget.cell(row=r, column=1).value
        budget = ws_budget.cell(row=r, column=2).value
        if cat and budget:
            budgets[cat] = budget

    actuals = {}
    for r in range(4, dm.last_row(ws_exp) + 1):
        cat = ws_exp.cell(row=r, column=3).value
        amt = ws_exp.cell(row=r, column=4).value
        if cat and amt and isinstance(amt, (int, float)):
            actuals[cat] = actuals.get(cat, 0) + amt

    print(f"\n  {'类别':<12} {'预算':>10} {'实际':>10} {'剩余':>10} {'状态':>8}")
    print("  " + "─" * 55)
    for cat in EXPENSE_CATEGORIES:
        b = budgets.get(cat, 0)
        a = actuals.get(cat, 0)
        if b > 0:
            remain = b - a
            status = "超支" if remain < 0 else ("紧张" if remain < b * 0.2 else "正常")
            icon = "⚠" if remain < 0 else "✓"
            print(f"  {cat:<12} {fmt(b):>10} {fmt(a):>10} {fmt(remain):>10} {icon}{status:>6}")

def func_29_budget_alert(dm: DataManager):
    """预算预警"""
    print_title("预算预警")
    ws_budget = dm.sheet("预算表", create=False)
    ws_exp = dm.sheet("支出记录", create=False)
    if not ws_budget or not ws_exp:
        print("  缺少数据"); return

    budgets = {}
    for r in range(4, dm.last_row(ws_budget) + 1):
        cat = ws_budget.cell(row=r, column=1).value
        budget = ws_budget.cell(row=r, column=2).value
        if cat and budget:
            budgets[cat] = budget

    actuals = {}
    for r in range(4, dm.last_row(ws_exp) + 1):
        cat = ws_exp.cell(row=r, column=3).value
        amt = ws_exp.cell(row=r, column=4).value
        if cat and amt and isinstance(amt, (int, float)):
            actuals[cat] = actuals.get(cat, 0) + amt

    alerts = 0
    for cat, b in budgets.items():
        a = actuals.get(cat, 0)
        if a > b * 0.8 and b > 0:
            alerts += 1
            pct = a / b * 100
            icon = "✗" if a > b else "⚠"
            print(f"  {icon} {cat}: 已用{pct:.0f}%（{fmt(a)}/{fmt(b)}）")

    if alerts == 0:
        print(f"  {c_ok('所有支出在预算范围内')}")
    else:
        print(f"\n  共 {alerts} 项预算预警")

# ============================================================================
# 功能30-33：数据工具
# ============================================================================

def func_30_query(dm: DataManager):
    """多条件查询"""
    print_title("多条件查询")
    target = input_choice("  查询目标", ["收入记录","支出记录"])
    if not target: return

    ws = dm.sheet(target, create=False)
    if not ws: print("  暂无数据"); return

    keyword = input("  关键词（模糊匹配，留空查全部）: ").strip()
    min_amt = input_num("  最小金额（留空不限）: ", 0)
    max_amt = input_num("  最大金额（留空不限）: ", 999999999)

    col_amt = 4  # 金额列
    col_name = 3  # 名称列（客户/类别）
    count = 0; total = 0

    print(f"\n  {'日期':<12} {'名称':<12} {'金额':>12}")
    print("  " + "─" * 40)
    for r in range(4, dm.last_row(ws) + 1):
        name_val = str(ws.cell(row=r, column=col_name).value or "")
        amt_val = ws.cell(row=r, column=col_amt).value
        dt_val = ws.cell(row=r, column=1).value

        if amt_val and isinstance(amt_val, (int, float)):
            if (not keyword or keyword in name_val) and min_amt <= amt_val <= max_amt:
                print(f"  {str(dt_val):<12} {name_val:<12} {fmt(amt_val):>12}")
                count += 1; total += amt_val

    print("  " + "─" * 40)
    print(f"  共 {count} 条，合计 {fmt(total)} 元")

def func_31_import_export(dm: DataManager):
    """数据导入导出"""
    print_title("数据导入导出")
    action = input_choice("  操作", ["导出CSV","导入CSV","导出Excel"], "导出CSV")
    if not action: return

    if "导出CSV" in action:
        target = input_choice("  工作表", ["收入记录","支出记录","应收应付","工资表"])
        if not target: return
        fp = os.path.join(DATA_DIR, f"{target}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv")
        ws = dm.sheet(target, create=False)
        if not ws: print("  表不存在"); return
        try:
            with open(fp, 'w', newline='', encoding='utf-8-sig') as f:
                writer = csv.writer(f)
                for row in ws.iter_rows(values_only=True):
                    writer.writerow([str(v) if v else "" for v in row])
            print(f"  ✓ 导出成功：{fp}")
        except Exception as e:
            print(f"  ✗ 导出失败：{e}")

    elif "导入CSV" in action:
        func_5_batch_import(dm)

    elif "导出Excel" in action:
        target = input_choice("  工作表", ["收入记录","支出记录","应收应付"])
        if not target: return
        fp = os.path.join(DATA_DIR, f"{target}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")
        ws = dm.sheet(target, create=False)
        if not ws: print("  表不存在"); return
        try:
            wb_new = Workbook()
            wb_new.remove(wb_new.active)
            ws_new = wb_new.create_sheet(title=target)
            for row in ws.iter_rows(values_only=True):
                ws_new.append(row)
            wb_new.save(fp)
            print(f"  ✓ 导出成功：{fp}")
        except Exception as e:
            print(f"  ✗ 导出失败：{e}")

def func_32_backup(dm: DataManager):
    """自动备份"""
    print_title("数据备份")
    if not os.path.exists(EXCEL_FILE):
        print("  无数据文件"); return

    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    bp = os.path.join(BACKUP_DIR, f"备份_{ts}.xlsx")
    shutil.copy2(EXCEL_FILE, bp)
    print(f"  ✓ 备份完成：{bp}")

    # 列出备份
    backups = sorted([f for f in os.listdir(BACKUP_DIR) if f.endswith('.xlsx')], reverse=True)
    print(f"\n  现有备份（{len(backups)}个）：")
    for b in backups[:5]:
        size = os.path.getsize(os.path.join(BACKUP_DIR, b)) / 1024
        print(f"    {b} ({size:.1f}KB)")

def func_33_data_repair(dm: DataManager):
    """数据修复"""
    print_title("数据修复工具")
    action = input_choice("  修复项目", ["修复日期格式","修复金额格式","清理空行","检查完整性","一键修复"], "一键修复")

    fixed = 0
    if action in ["修复日期格式", "一键修复"]:
        for name in ["收入记录", "支出记录"]:
            ws = dm.sheet(name, create=False)
            if ws:
                for r in range(4, dm.last_row(ws) + 1):
                    v = ws.cell(row=r, column=1).value
                    if v:
                        try:
                            ws.cell(row=r, column=1).value = str(v)[:10]
                            fixed += 1
                        except: pass
        print(f"  ✓ 日期格式修复 {fixed} 条")

    if action in ["修复金额格式", "一键修复"]:
        for name in ["收入记录", "支出记录"]:
            ws = dm.sheet(name, create=False)
            if ws:
                for r in range(4, dm.last_row(ws) + 1):
                    v = ws.cell(row=r, column=4).value
                    if v and isinstance(v, (int, float)):
                        ws.cell(row=r, column=4).number_format = '#,##0.00'
                        fixed += 1
        print(f"  ✓ 金额格式修复 {fixed} 条")

    if action in ["清理空行", "一键修复"]:
        for name in ["收入记录", "支出记录"]:
            ws = dm.sheet(name, create=False)
            if ws:
                deleted = 0
                for r in range(dm.last_row(ws), 3, -1):
                    if ws.cell(row=r, column=1).value is None and ws.cell(row=r, column=3).value is None:
                        ws.delete_rows(r)
                        deleted += 1
                print(f"  ✓ {name} 清理空行 {deleted} 行")

    if action in ["检查完整性", "一键修复"]:
        sheets = ["收入记录", "支出记录", "应收应付", "利润分析表", "工资表"]
        for s in sheets:
            ws = dm.sheet(s, create=False)
            status = "✓" if ws else "✗ 缺失"
            print(f"  {status} {s}")

    if action != "检查完整性":
        dm.save()

# ============================================================================
# 功能34-36：智能提醒
# ============================================================================

def func_34_ar_alert(dm: DataManager):
    """应收逾期提醒"""
    print_title("应收逾期提醒")
    ws = dm.sheet("应收应付", create=False)
    if not ws or dm.last_row(ws) < 4:
        print("  暂无数据"); return

    alerts = 0; total = 0
    for r in range(4, dm.last_row(ws) + 1):
        cust = ws.cell(row=r, column=1).value
        amt = ws.cell(row=r, column=6).value
        if amt and isinstance(amt, (int, float)) and amt > 0:
            alerts += 1; total += amt
            icon = "⚠" if amt > 20000 else ("△" if amt > 10000 else "✓")
            print(f"  {icon} {cust}: {fmt(amt)} 元")

    if alerts == 0:
        print(f"  {c_ok('暂无应收需要关注')}")
    else:
        print(f"\n  共 {alerts} 条应收，合计 {fmt(total)} 元")

def func_35_inventory_alert(dm: DataManager):
    """库存预警"""
    print_title("库存预警")
    ws = dm.sheet("材料进销存台账", create=False)
    if not ws or dm.last_row(ws) < 4:
        print("  暂无数据"); return

    inventory = {}
    for r in range(4, dm.last_row(ws) + 1):
        mat = ws.cell(row=r, column=2).value
        qty_in = ws.cell(row=r, column=4).value or 0
        qty_out = ws.cell(row=r, column=7).value or 0
        if mat:
            inventory[mat] = inventory.get(mat, 0) + qty_in - qty_out

    alerts = 0
    for mat, qty in sorted(inventory.items()):
        if qty < 10:
            alerts += 1
            icon = "✗" if qty <= 0 else "⚠"
            print(f"  {icon} {mat}: 库存 {qty}（{'缺货' if qty <= 0 else '不足'}）")

    if alerts == 0:
        print(f"  {c_ok('所有材料库存充足')}")
    else:
        print(f"\n  共 {alerts} 种材料需要采购")

def func_36_profit_alert(dm: DataManager):
    """利润预警"""
    print_title("利润预警")

    ws_inc = dm.sheet("收入记录", create=False)
    ws_exp = dm.sheet("支出记录", create=False)
    income = expense = 0
    if ws_inc:
        for r in range(4, dm.last_row(ws_inc) + 1):
            v = ws_inc.cell(row=r, column=4).value
            if v and isinstance(v, (int, float)): income += v
    if ws_exp:
        for r in range(4, dm.last_row(ws_exp) + 1):
            v = ws_exp.cell(row=r, column=4).value
            if v and isinstance(v, (int, float)): expense += v

    profit = income - expense
    rate = profit / income * 100 if income > 0 else 0

    print(f"\n  收入：{fmt(income)} 元")
    print(f"  支出：{fmt(expense)} 元")
    print(f"  利润：{fmt(profit)} 元")
    print(f"  利润率：{rate:.1f}%")

    if profit < 0:
        print("\n  ✗ 严重警告：当前亏损！请立即控制成本！")
    elif rate < 10:
        print("\n  ⚠ 警告：利润率偏低（<10%）")
    elif rate > 30:
        print("\n  ✓ 利润率良好（>30%）")
    else:
        print("\n  ✓ 经营状况正常")

# ============================================================================
# 功能37-41：常用模板
# ============================================================================

def func_37_salary_template(dm: DataManager):
    """工资表模板 - 查看/清空工资表"""
    print_title("工资表管理")
    ws = dm.sheet("工资表")
    last = dm.last_row(ws, 2)
    if last < 5:
        print("  工资表为空，请通过菜单[17]录入工资")
    else:
        print(f"  当前工资记录：{last - 4} 条")
        print(f"  {'序号':<4} {'姓名':<8} {'应发合计':>10} {'实发合计':>10}")
        print("  " + "─" * 36)
        for r in range(5, last + 1):
            name = ws.cell(row=r, column=2).value or ""
            gross = ws.cell(row=r, column=6).value or 0
            net = ws.cell(row=r, column=10).value or 0
            print(f"  {r-4:<4} {name:<8} {gross:>10,.2f} {net:>10,.2f}")
    action = input_choice("  操作", ["查看完毕", "清空数据"], "查看完毕")
    if "清空" in action:
        for r in range(5, last + 1):
            for c in range(1, 12):
                ws.cell(row=r, column=c).value = None
        dm.save()
        print(f"  {c_ok('工资表已清空')}")

def func_38_invoice_template(dm: DataManager):
    """发票登记 - 快速登记发票"""
    print_title("发票登记")
    ws = dm.sheet("发票登记表")
    dt = input_date("  开票日期")
    no = input("  发票号码: ").strip()
    if not no:
        print("  " + c_err('发票号码不能为空')); return
    itype = input_choice("  发票类型", ["增值税专用发票", "增值税普通发票", "收据"], "增值税普通发票")
    customer = input("  客户名称: ").strip()
    amount = input_num("  金额（不含税）: ")
    tax_amt = round(amount * 0.03, 2)
    total = round(amount + tax_amt, 2)
    status = input_choice("  发票状态", ["已开具", "已作废", "已红冲"], "已开具")
    remark = input("  备注: ").strip()
    row = dm.next_row(ws)
    dm.write_row(ws, row, [dt, no, itype, customer, amount, tax_amt, total, status, remark])
    dm.save()
    print(f"  ✓ 发票已登记，价税合计：{fmt(total)} 元")

def func_39_bank_template(dm: DataManager):
    """银行对账 - 录入银行流水"""
    print_title("银行对账录入")
    ws = dm.sheet("银行对账表")
    dt = input_date("  日期")
    desc = input("  摘要: ").strip()
    bank_in = input_num("  银行收入（默认0）: ", 0)
    bank_out = input_num("  银行支出（默认0）: ", 0)
    remark = input("  备注: ").strip()
    row = dm.next_row(ws)
    dm.write_row(ws, row, [dt, desc, bank_in if bank_in > 0 else "", bank_out if bank_out > 0 else "",
                           "", "", "", "", "", remark])
    dm.save()
    print(f"  ✓ 银行流水已录入第{row}行")

def func_40_material_template(dm: DataManager):
    """材料台账 - 快速查看库存"""
    print_title("材料库存概览")
    ws = dm.sheet("材料进销存台账")
    last = dm.last_row(ws, 2)
    if last < 4:
        print("  材料台账为空")
        return
    inventory = {}
    for r in range(4, last + 1):
        mat = ws.cell(row=r, column=2).value
        qty_in = ws.cell(row=r, column=4).value or 0
        qty_out = ws.cell(row=r, column=7).value or 0
        if mat:
            inventory[mat] = inventory.get(mat, 0) + qty_in - qty_out
    print(f"  {'材料名称':<12} {'当前库存':>10}")
    print("  " + "─" * 26)
    for mat, qty in sorted(inventory.items()):
        flag = " ⚠缺货" if qty <= 0 else ""
        print(f"  {mat:<12} {qty:>10,.2f}{flag}")

def func_41_monthly_template(dm: DataManager):
    """月度报表 - 快速生成月度经营报表"""
    print_title("月度经营报表")
    month = input("  请输入月份（如2024-03）: ").strip()
    if not month or len(month) != 7:
        print("  " + c_err('月份格式不正确')); return
    ws = dm.sheet("月度经营报表")
    # 检查是否已有该月数据
    found = False
    for r in range(4, dm.last_row(ws) + 1):
        if ws.cell(row=r, column=1).value == month:
            found = True
            break
    if found:
        print(f"  {month} 的报表已存在")
        return
    # 从收入/支出表汇总
    inc_ws = dm.sheet("收入记录")
    exp_ws = dm.sheet("支出记录")
    total_inc = 0
    total_exp = 0
    for r in range(4, dm.last_row(inc_ws) + 1):
        if str(inc_ws.cell(row=r, column=1).value or "").startswith(month):
            total_inc += inc_ws.cell(row=r, column=4).value or 0
    for r in range(4, dm.last_row(exp_ws) + 1):
        if str(exp_ws.cell(row=r, column=1).value or "").startswith(month):
            total_exp += exp_ws.cell(row=r, column=4).value or 0
    row = dm.next_row(ws)
    profit = total_inc - total_exp
    dm.write_row(ws, row, [month, fmt(total_inc), fmt(total_exp), fmt(profit),
                           f"{profit/total_inc*100:.1f}%" if total_inc > 0 else "N/A"])
    dm.save()
    print(f"  ✓ {month} 月度报表已生成")
    print(f"    收入：{fmt(total_inc)} | 支出：{fmt(total_exp)} | 利润：{fmt(profit)}")

# ============================================================================
# 功能42-45：系统
# ============================================================================

def func_42_sample_data(dm: DataManager):
    """生成模拟数据"""
    amounts = [15000, 12000, 18000, 8500, 22000]
    now = date.today()
    months = [(now.month - i - 1) % 12 + 1 for i in range(2, -1, -1)]
    years = [now.year - (1 if m > now.month else 0) for m in months]

    print("\n" + "─" * 40)
    print(f"  生成模拟数据（{years[0]}年{months[0]}月-{years[-1]}年{months[-1]}月）")
    print_divider()
    confirm = input("  确认生成？将添加模拟数据（y/n）: ").strip().lower()
    if confirm != 'y':
        print("  已取消"); return
    for mi, month in enumerate(months):
        yr = years[mi]
        for i, cust in enumerate(SAMPLE_CUSTOMERS):
            for j in range(2):
                dt = f"{yr}-{month:02d}-{5+j*10:02d}"
                amt = amounts[i] * (0.8 + random.random() * 0.4)
                ws = dm.sheet("收入记录")
                row = dm.next_row(ws)
                voucher = f"收-{month:02d}-{row-3:03d}"
                tax_type = "含税" if i % 2 == 0 else "不含税"
                rate = 0.03
                if tax_type == "含税":
                    tax = round(amt * rate / (1 + rate), 2)
                    no_tax = round(amt / (1 + rate), 2)
                else:
                    tax = round(amt * rate, 2)
                    no_tax = amt
                dm.write_row(ws, row, [dt, voucher, cust, round(amt, 2),
                    random.choice(["银行转账","微信"]), tax_type, rate, tax, no_tax,
                    random.choice(["已开票","未开票"]), "正常", "", ""])

        for cat, base_amt in [("厂租",8000),("电费",4000),("水费",400),("工资",15000),
                                ("社保费",2500),("化工",3000),("染料封闭剂",2000),
                                ("包装材料",800),("货车运输费",1500),("外发加工费",2500),
                                ("税金",500),("业务费",500),("办公用品费",200),
                                ("维修费",500),("挂具费",400)]:
            dt = f"{yr}-{month:02d}-{random.randint(5,25):02d}"
            ws = dm.sheet("支出记录")
            row = dm.next_row(ws)
            voucher = f"支-{month:02d}-{row-3:03d}"
            amt = base_amt * (0.9 + random.random() * 0.2)
            dm.write_row(ws, row, [dt, voucher, cat, round(amt, 2), "", "银行转账", ""])

    # 应收应付
    ws = dm.sheet("应收应付")
    for i, cust in enumerate(SAMPLE_CUSTOMERS):
        row = 4 + i
        init_ar = amounts[i] * 0.3
        add_ar = amounts[i] * 3
        recv = amounts[i] * 2.5
        dm.write_row(ws, row, [cust, init_ar, add_ar, recv, 0, init_ar + add_ar - recv, ""])

    # 工资
    ws = dm.sheet("工资表")
    for i, (name, base) in enumerate([("张三",3500),("李四",4000),("王五",4500)]):
        row = 5 + i
        ot = round(200 + random.random() * 500, 2)
        bonus = round(100 + random.random() * 300, 2)
        gross = base + ot + bonus
        si = gross * 0.1
        tax = calc_income_tax(gross - si - 5000)
        net = gross - si - tax
        dm.write_row(ws, row, [i+1, name, base, ot, bonus, gross, si, 0, round(tax,2), round(net,2), "银行转账"])

    # 材料
    ws = dm.sheet("材料进销存台账")
    prices = [180, 150, 45, 38, 42, 280, 120, 95]
    specs = ["25kg/袋","25kg/袋","500ml/瓶","500ml/瓶","500ml/瓶","20kg/桶","5kg/桶","5kg/桶"]
    row = 4
    for mi, month in enumerate(months):
        yr = years[mi]
        for mi2, mat in enumerate(MATERIALS):
            qty_in = 10 + random.randint(0, 5)
            dm.write_row(ws, row, [f"{yr}-{month:02d}-05", mat, specs[mi2], qty_in, prices[mi2], qty_in*prices[mi2]])
            row += 1
            qty_out = 8 + random.randint(0, 4)
            dm.write_row(ws, row, [f"{yr}-{month:02d}-15", mat, specs[mi2], None, None, None, qty_out, prices[mi2], qty_out*prices[mi2]])
            row += 1

    dm.save()
    print(f"  {c_ok('模拟数据生成完成！')}")
    print("    收入：30条 | 支出：45条 | 应收：5条 | 工资：3人 | 材料：48条")

def func_43_settings(dm: DataManager):
    """系统设置"""
    print_title("系统设置")
    print(f"  数据目录：{DATA_DIR}")
    print(f"  备份目录：{BACKUP_DIR}")
    print(f"  数据文件：{EXCEL_FILE}")
    print(f"  文件大小：{os.path.getsize(EXCEL_FILE)/1024:.1f} KB" if os.path.exists(EXCEL_FILE) else "  文件不存在")

def func_44_help(dm: DataManager):
    """使用帮助"""
    print(f"\n  {VERSION}")
    print("  ─" * 25)
    print("  运行：python 氧化加工厂财务系统.py")
    print("  依赖：pip install openpyxl")
    print("  数据：财务数据/财务数据.xlsx")
    print("  备份：备份/ 目录")
    print("  功能：45个财务管理功能")
    print("  模板：5个常用模板")

def func_45_about(dm: DataManager):
    """关于系统"""
    print(f"\n  ╔══════════════════════════════════╗")
    print(f"  ║  小型氧化加工厂财务管理系统      ║")
    print(f"  ║  {VERSION:<28s}║")
    print(f"  ║  更新：2026年5月                  ║")
    print(f"  ║  功能：45个 + 5个模板            ║")
    print(f"  ║  语言：Python 3.8+               ║")
    print(f"  ║  依赖：openpyxl                  ║")
    print(f"  ╚══════════════════════════════════╝")

# ============================================================================
# 主菜单
# ============================================================================

# 功能映射表
FUNC_MAP = {
    "1": ("录入收入", func_1_add_income),
    "2": ("录入支出", func_2_add_expense),
    "3": ("收入汇总", func_3_income_summary),
    "4": ("支出汇总", func_4_expense_summary),
    "5": ("批量导入", func_5_batch_import),
    "6": ("利润报表", func_6_profit_report),
    "7": ("资产负债表", func_7_balance_sheet),
    "8": ("月度报表", func_8_monthly_report),
    "9": ("客户对账单", func_9_customer_statement),
    "10": ("多期对比", func_10_multi_period),
    "11": ("增值税计算", func_11_vat),
    "12": ("所得税计算", func_12_income_tax),
    "13": ("六税两费", func_13_six_taxes),
    "14": ("残保金", func_14_disabled_fund),
    "15": ("工会经费", func_15_union_fund),
    "16": ("税务汇总", func_16_tax_summary),
    "17": ("工资计算", func_17_payroll),
    "18": ("个税计算", func_18_tax_calc),
    "19": ("社保明细", func_19_social_insurance),
    "20": ("固定资产", func_20_fixed_assets),
    "21": ("折旧计算", func_21_depreciation),
    "22": ("低值易耗品", func_22_low_value),
    "23": ("成本核算", func_23_cost_accounting),
    "24": ("差异分析", func_24_variance_analysis),
    "25": ("应收账龄", func_25_ar_aging),
    "26": ("应付账龄", func_26_ap_aging),
    "27": ("预算编制", func_27_budget),
    "28": ("预算控制", func_28_budget_control),
    "29": ("预算预警", func_29_budget_alert),
    "30": ("多条件查询", func_30_query),
    "31": ("导入导出", func_31_import_export),
    "32": ("自动备份", func_32_backup),
    "33": ("数据修复", func_33_data_repair),
    "34": ("应收逾期", func_34_ar_alert),
    "35": ("库存预警", func_35_inventory_alert),
    "36": ("利润预警", func_36_profit_alert),
    "37": ("工资表模板", func_37_salary_template),
    "38": ("发票登记", func_38_invoice_template),
    "39": ("银行对账", func_39_bank_template),
    "40": ("材料台账", func_40_material_template),
    "41": ("月度报表", func_41_monthly_template),
    "42": ("模拟数据", func_42_sample_data),
    "43": ("系统设置", func_43_settings),
    "44": ("使用帮助", func_44_help),
    "45": ("关于系统", func_45_about),
}

def show_dashboard(dm: DataManager):
    """启动仪表盘 - 显示数据概览"""
    clear()
    # 统计数据
    ws_inc = dm.sheet("收入记录", create=False)
    ws_exp = dm.sheet("支出记录", create=False)
    ws_ar = dm.sheet("应收应付", create=False)
    ws_sal = dm.sheet("工资表", create=False)
    ws_mat = dm.sheet("材料进销存台账", create=False)

    inc_count = max(0, dm.last_row(ws_inc) - 3) if ws_inc else 0
    exp_count = max(0, dm.last_row(ws_exp) - 3) if ws_exp else 0
    ar_count = max(0, dm.last_row(ws_ar) - 3) if ws_ar else 0
    sal_count = max(0, dm.last_row(ws_sal) - 4) if ws_sal else 0
    mat_count = max(0, dm.last_row(ws_mat) - 3) if ws_mat else 0

    total_inc = 0
    if ws_inc:
        for r in range(4, dm.last_row(ws_inc) + 1):
            v = ws_inc.cell(row=r, column=4).value
            if v and isinstance(v, (int, float)):
                total_inc += v

    total_exp = 0
    if ws_exp:
        for r in range(4, dm.last_row(ws_exp) + 1):
            v = ws_exp.cell(row=r, column=4).value
            if v and isinstance(v, (int, float)):
                total_exp += v

    profit = total_inc - total_exp
    profit_rate = (profit / total_inc * 100) if total_inc > 0 else 0

    total_ar = 0
    if ws_ar:
        for r in range(4, dm.last_row(ws_ar) + 1):
            v = ws_ar.cell(row=r, column=6).value
            if v and isinstance(v, (int, float)):
                total_ar += v

    # 打印仪表盘
    W = 50
    print(cr(f"╔{'═' * W}╗", C.B))
    print(cr(f"║{'小型氧化加工厂财务管理系统':^{W}}║", C.BD + C.B))
    print(cr(f"║{VERSION:^{W}}║", C.D))
    print(cr(f"╠{'═' * W}╣", C.B))
    print(cr(f"║  {date.today().strftime('%Y年%m月%d日 %A'):^{W-2}}║", C.W))
    print(cr(f"╠{'═' * W}╣", C.B))

    # 数据概览
    print(cr(f"║  {'📊 数据概览':^{W-4}}║", C.BD + C.M))
    print(cr(f"╟{'─' * W}╢", C.B))
    data_items = [
        ("收入记录", f"{inc_count} 条", c_num),
        ("支出记录", f"{exp_count} 条", c_num),
        ("应收应付", f"{ar_count} 条", c_num),
        ("工资记录", f"{sal_count} 人", c_num),
        ("材料台账", f"{mat_count} 条", c_num),
    ]
    for label, val, color_fn in data_items:
        print(f"║  {label:8s} {color_fn(val):>{W-16}}║")

    print(cr(f"╟{'─' * W}╢", C.B))

    # 财务摘要
    print(cr(f"║  {'💰 财务摘要':^{W-4}}║", C.BD + C.M))
    print(cr(f"╟{'─' * W}╢", C.B))
    fin_items = [
        ("总收入", f"¥ {fmt(total_inc)}", c_money),
        ("总支出", f"¥ {fmt(total_exp)}", c_money),
        ("净利润", f"¥ {fmt(profit)}", c_ok if profit >= 0 else c_err),
        ("利润率", f"{profit_rate:.1f}%", c_ok if profit_rate > 10 else c_warn),
        ("应收款", f"¥ {fmt(total_ar)}", c_warn if total_ar > 0 else c_ok),
    ]
    for label, val, color_fn in fin_items:
        print(f"║  {label:8s} {color_fn(val):>{W-16}}║")

    print(cr(f"╚{'═' * W}╝", C.B))
    print()

def show_menu():
    """显示主菜单"""
    clear()
    W = 50
    print(cr(f"╔{'═' * W}╗", C.B))
    print(cr(f"║{'主菜单':^{W}}║", C.BD + C.B))
    print(cr(f"╠{'═' * W}╣", C.B))

    sections = [
        ("一、日常记账", [("1","录入收入"),("2","录入支出"),("3","收入汇总"),("4","支出汇总"),("5","批量导入")]),
        ("二、报表生成", [("6","利润报表"),("7","资产负债"),("8","月度报表"),("9","客户对账"),("10","多期对比")]),
        ("三、税务管理", [("11","增值税"),("12","所得税"),("13","六税两费"),("14","残保金"),("15","工会经费"),("16","税务汇总")]),
        ("四、工资社保", [("17","工资计算"),("18","个税计算"),("19","社保明细")]),
        ("五、资产管理", [("20","固定资产"),("21","折旧计算"),("22","低值易耗品")]),
        ("六、成本分析", [("23","成本核算"),("24","差异分析"),("25","应收账龄"),("26","应付账龄")]),
        ("七、预算管理", [("27","预算编制"),("28","预算控制"),("29","预算预警")]),
        ("八、数据工具", [("30","条件查询"),("31","导入导出"),("32","自动备份"),("33","数据修复")]),
        ("九、智能提醒", [("34","应收逾期"),("35","库存预警"),("36","利润预警")]),
        ("十、常用模板", [("37","工资表"),("38","发票登记"),("39","银行对账"),("40","材料台账"),("41","月度报表")]),
        ("系统", [("42","模拟数据"),("43","系统设置"),("44","使用帮助"),("45","关于系统")]),
    ]

    # 两列布局
    pairs = []
    for i in range(0, len(sections) - 1, 2):
        pairs.append((sections[i], sections[i + 1]))
    if len(sections) % 2 == 1:
        pairs.append((sections[-1], None))

    for left, right in pairs:
        # 区块标题行
        l_title = cr(f"◆ {left[0]} ◆", C.BD + C.M)
        if right:
            r_title = cr(f"◆ {right[0]} ◆", C.BD + C.M)
            print(f"║  {l_title:<22s}  {r_title:<22s}║")
        else:
            print(f"║  {l_title:<22s}{' ' * 24}║")

        # 功能项行
        l_items = left[1]
        r_items = right[1] if right else []
        max_rows = max(len(l_items), len(r_items))
        for ri in range(max_rows):
            l_str = ""
            if ri < len(l_items):
                num, name = l_items[ri]
                l_str = cr(f"[{num:>2s}]", C.Y) + f" {name}"
            r_str = ""
            if ri < len(r_items):
                num, name = r_items[ri]
                r_str = cr(f"[{num:>2s}]", C.Y) + f" {name}"
            print(f"║  {l_str:<22s}  {r_str:<22s}║")
        print(cr(f"╟{'─' * W}╢", C.B))

    print(cr(f"║  {cr('[ 0]', C.R)} 退出系统{' ' * 36}║", C.W))
    print(cr(f"╚{'═' * W}╝", C.B))

def show_goodbye(dm: DataManager):
    """退出画面"""
    clear()
    W = 50
    # 统计
    ws_inc = dm.sheet("收入记录", create=False)
    ws_exp = dm.sheet("支出记录", create=False)
    total_inc = 0
    total_exp = 0
    if ws_inc:
        for r in range(4, dm.last_row(ws_inc) + 1):
            v = ws_inc.cell(row=r, column=4).value
            if v and isinstance(v, (int, float)): total_inc += v
    if ws_exp:
        for r in range(4, dm.last_row(ws_exp) + 1):
            v = ws_exp.cell(row=r, column=4).value
            if v and isinstance(v, (int, float)): total_exp += v
    profit = total_inc - total_exp

    print()
    print(cr(f"╔{'═' * W}╗", C.B))
    print(cr(f"║{'感谢使用小型氧化加工厂财务管理系统':^{W}}║", C.BD + C.B))
    print(cr(f"╠{'═' * W}╣", C.B))
    print(cr(f"║{'本次会话数据摘要':^{W}}║", C.M))
    print(cr(f"╟{'─' * W}╢", C.B))
    print(f"║  累计收入：{c_money(f'¥ {fmt(total_inc)}'):>{W-14}}║")
    print(f"║  累计支出：{c_money(f'¥ {fmt(total_exp)}'):>{W-14}}║")
    profit_color = c_ok if profit >= 0 else c_err
    print(f"║  净利润：  {profit_color(f'¥ {fmt(profit)}'):>{W-14}}║")
    print(cr(f"╟{'─' * W}╢", C.B))
    print(cr(f"║{'数据已自动保存':^{W}}║", C.G))
    print(cr(f"║{VERSION:^{W}}║", C.D))
    print(cr(f"╚{'═' * W}╝", C.B))
    print()

def main():
    """主函数"""
    # 初始化
    dm = DataManager()
    if not os.path.exists(EXCEL_FILE):
        print("首次运行，正在初始化系统...")
        init_all_sheets(dm)
        pause()

    # 主循环
    while True:
        show_menu()
        choice = input(cr("请选择功能编号: ", C.Y)).strip()

        if choice == "0":
            show_goodbye(dm)
            break
        elif choice == "":
            show_dashboard(dm)
            pause()
            continue
        elif choice in FUNC_MAP:
            name, func = FUNC_MAP[choice]
            try:
                func(dm)
            except Exception as e:
                print(f"\n  {c_err(f'执行出错：{e}')}")
            pause()
        else:
            print(f"\n  {c_warn('无效编号，请输入 0-45')}")
            pause()

if __name__ == '__main__':
    try:
        main()
    except KeyboardInterrupt:
        print(f"\n\n  {c_dim('程序已退出')}")
